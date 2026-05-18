
import copy
import glob
import hashlib
import json
import logging
import os
import re
import time
import zipfile
import xml.etree.ElementTree as ET
from collections import defaultdict, deque
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from difflib import SequenceMatcher
from threading import Lock, Semaphore
from typing import Any, Optional

try:
    import pandas as pd
except Exception:
    pd = None
from urllib import error as urlerror
from urllib import request as urlrequest


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 1) 全局配置
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("TARA-V10")


MODEL_NAME = os.getenv("TARA_MODEL", "qwen3.5-122b-a10b")
API_KEY = os.getenv("TARA_API_KEY", "sk-3458388d4845414387fe2e10bc8a9ee2")
API_BASE = os.getenv("TARA_API_BASE", "https://dashscope.aliyuncs.com/compatible-mode/v1")

if not API_KEY:
    logger.warning("未检测到环境变量 TARA_API_KEY。")

REQUEST_TIMEOUT = int(os.getenv("TARA_REQUEST_TIMEOUT", "90"))
MAX_TOKENS = int(os.getenv("TARA_MAX_TOKENS", "4096"))
TEMPERATURE = float(os.getenv("TARA_TEMPERATURE", "0.1"))

MAX_WORKERS = int(os.getenv("TARA_MAX_WORKERS", "6"))
MAX_CONCURRENT_LLM = int(os.getenv("TARA_MAX_CONCURRENT_LLM", "6"))
CALL_INTERVAL = float(os.getenv("TARA_CALL_INTERVAL", "0.4"))

BATCH_SIZE_INFLUENCE = int(os.getenv("TARA_BATCH_SIZE_INFLUENCE", "8"))
BATCH_SIZE_ATTACK = int(os.getenv("TARA_BATCH_SIZE_ATTACK", "8"))

# 可选：attack_potential / attack_vector / cvss
FEASIBILITY_METHOD = "cvss"

OUTPUT_DIR = r"D:\Jupyter profile\汽车信息安全风险评估\outputs\V13"
os.makedirs(OUTPUT_DIR, exist_ok=True)

ENABLE_RAG = os.getenv("TARA_ENABLE_RAG", "1").strip() not in {"0", "false", "False"}
RAG_BASE_DIR = os.getenv("TARA_RAG_BASE_DIR", r"D:\Jupyter profile\汽车信息安全风险评估\data\rag")
RAG_DIRS = {
    "tara_reports": os.path.join(RAG_BASE_DIR, "tara_reports"),
    "regulations": os.path.join(RAG_BASE_DIR, "regulations"),
    "attack_databases": os.path.join(RAG_BASE_DIR, "attack_databases"),
}
RAG_TOP_K = int(os.getenv("TARA_RAG_TOP_K", "5"))
RAG_MAX_CONTEXT_CHARS = int(os.getenv("TARA_RAG_MAX_CHARS", "1800"))
RAG_BM25_K1 = float(os.getenv("TARA_BM25_K1", "1.5"))   # BM25 词频饱和系数
RAG_BM25_B = float(os.getenv("TARA_BM25_B", "0.75"))    # BM25 文档长度归一化

llm_semaphore = Semaphore(MAX_CONCURRENT_LLM)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 1.5) 计时工具
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class Timer:
    """分阶段计时器：with 块自动记录耗时并日志输出。"""

    def __init__(self, stage_name: str):
        self.stage_name = stage_name
        self.start = 0.0

    def __enter__(self):
        self.start = time.perf_counter()
        logger.info("═══ [计时] %s 开始 ═══", self.stage_name)
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        elapsed = time.perf_counter() - self.start
        if exc_type is None:
            logger.info("═══ [计时] %s 完成，耗时 %.2f 秒 (%.2f 分钟) ═══", self.stage_name, elapsed, elapsed / 60)
        else:
            logger.error("═══ [计时] %s 异常退出，耗时 %.2f 秒，错误: %s ═══", self.stage_name, elapsed, exc_val)


# 阶段累计计时器（线程安全）
_phase_times: dict[str, float] = {"damage": 0.0, "threat": 0.0, "attack_path": 0.0, "impact": 0.0, "feasibility": 0.0}
_phase_lock = Lock()


def _add_phase_time(phase: str, elapsed: float) -> None:
    with _phase_lock:
        _phase_times[phase] = _phase_times.get(phase, 0.0) + elapsed


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 2) 业务常量
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

SECURITY_ATTRIBUTES_MAP: dict[str, list[str]] = {
    "数据": ["完整性", "机密性", "可用性"],
    "信号": ["完整性", "机密性", "真实性", "可用性"],
    "部件": ["完整性", "机密性", "真实性", "不可抵赖性", "权限属性", "可用性"],
    "接口": ["完整性", "机密性", "真实性", "可用性"],
}

ATTRIBUTE_TO_THREAT: dict[str, str] = {
    "完整性": "篡改",
    "机密性": "信息泄露",
    "可用性": "拒绝服务",
    "真实性": "欺骗",
    "不可抵赖性": "抵赖",
    "权限属性": "越权",
}

# Attack Potential 五维取值规范（ISO/SAE 21434 常见量表）
AP_ALLOWED_VALUES = {
    "Exposure_time": [0, 1, 4, 17, 19],
    "Professional_experience": [0, 3, 6, 9],
    "Required_information": [0, 3, 7, 11],
    "Opportunity_window": [0, 1, 4, 10],
    "Required_equipment": [0, 4, 7, 9],
}

# CVSS 取值
CVSS_V = {
    "network": 0.85,
    "adjacent": 0.62,
    "local": 0.55,
    "physical": 0.20,
}
CVSS_C = {"low": 0.77, "high": 0.44}
CVSS_P = {"none": 0.85, "low": 0.62, "high": 0.27}
CVSS_U = {"none": 0.85, "required": 0.62}

# 攻击向量 -> 可行性等级映射（ISO 21434 G.4）
ATTACK_VECTOR_TO_LEVEL = {
    "network": "High",
    "adjacent": "Medium",
    "local": "Low",
    "physical": "Very Low",
}

ATTACK_VECTOR_SCORE = {
    "network": 0.85,
    "adjacent": 0.62,
    "local": 0.55,
    "physical": 0.20,
}

RISK_MATRIX = {
    (4, "High"): 5, (4, "Medium"): 5, (4, "Low"): 4, (4, "Very Low"): 4,
    (3, "High"): 5, (3, "Medium"): 4, (3, "Low"): 4, (3, "Very Low"): 3,
    (2, "High"): 4, (2, "Medium"): 4, (2, "Low"): 3, (2, "Very Low"): 2,
    (1, "High"): 3, (1, "Medium"): 2, (1, "Low"): 2, (1, "Very Low"): 1,
}


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 3) 语义缓存
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class SemanticCache:
    """轻量语义缓存：规范化文本 + 相似度匹配。"""

    def __init__(self, similarity_threshold: float = 0.985, max_bucket_size: int = 1000):
        self.similarity_threshold = similarity_threshold
        self.max_bucket_size = max_bucket_size
        self._store: dict[str, list[dict]] = defaultdict(list)
        self._lock = Lock()

    @staticmethod
    def _normalize(obj: Any) -> str:
        if isinstance(obj, (dict, list, tuple)):
            text = json.dumps(obj, ensure_ascii=False, sort_keys=True)
        else:
            text = str(obj)
        text = text.lower().strip()
        text = re.sub(r"\s+", "", text)
        text = text.replace("：", ":")
        return text

    def get(self, bucket: str, key_obj: Any) -> Optional[Any]:
        signature = self._normalize(key_obj)
        key_hash = hashlib.sha256(signature.encode("utf-8")).hexdigest()

        with self._lock:
            items = self._store.get(bucket, [])
            for item in items:
                if item["hash"] == key_hash:
                    return copy.deepcopy(item["value"])
                ratio = SequenceMatcher(None, signature, item["signature"]).ratio()
                if ratio >= self.similarity_threshold:
                    return copy.deepcopy(item["value"])
        return None

    def set(self, bucket: str, key_obj: Any, value: Any) -> None:
        signature = self._normalize(key_obj)
        key_hash = hashlib.sha256(signature.encode("utf-8")).hexdigest()

        with self._lock:
            items = self._store[bucket]
            for item in items:
                if item["hash"] == key_hash:
                    item["value"] = copy.deepcopy(value)
                    return
            items.append({"hash": key_hash, "signature": signature, "value": copy.deepcopy(value)})
            if len(items) > self.max_bucket_size:
                del items[0 : len(items) - self.max_bucket_size]


semantic_cache = SemanticCache()

# 攻击路径缓存：相同 asset_name 的攻击路径相同，跨功能/属性复用
_attack_path_cache: dict[str, list[str]] = {}
_attack_path_cache_lock = Lock()


def _get_attack_path_cache(asset_name: str) -> list[str] | None:
    with _attack_path_cache_lock:
        return copy.deepcopy(_attack_path_cache.get(asset_name))


def _set_attack_path_cache(asset_name: str, paths: list[str]) -> None:
    with _attack_path_cache_lock:
        _attack_path_cache[asset_name] = copy.deepcopy(paths)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 3.5) BM25 RAG（jieba 分词 + BM25 Okapi 关键词匹配）
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

import math


class BM25RAGKnowledgeBase:
    """基于 BM25 关键词匹配的 RAG，无 GPU 依赖，秒级启动。

    数据源为 JSONL 预切块文件（每条记录一个语义块），
    辅以 PDF/Excel/TXT/MD/JSON 源文件。
    使用 jieba 分词 + BM25 Okapi 评分，按 ``tara_phase`` 元数据实现阶段感知检索。
    """

    def __init__(self, rag_dirs: dict[str, str], k1: float = 1.5, b: float = 0.75):
        self.rag_dirs = rag_dirs
        self.k1 = k1
        self.b = b
        self._lock = Lock()
        self._loaded = False
        self.records: list[dict] = []      # [{text, category, path, library_name, ...}]
        self._bm25: "BM25RAGKnowledgeBase._BM25Okapi | None" = None

    # ── 同义词扩展 ──────────────────────────────────

    # 汽车网络安全领域同义词表，用于查询扩展提升召回
    SYNONYM_MAP: dict[str, str] = {
        "机密性": "保密性 隐私",
        "完整性": "篡改 数据篡改 消息篡改",
        "可用性": "拒绝服务 功能丧失 停用",
        "真实性": "身份验证 认证 伪造",
        "ECU": "电子控制单元 控制器",
        "CAN": "控制器局域网 总线 CAN总线",
        "T-Box": "远程信息处理 车载通信终端",
        "OBD": "车载诊断 诊断接口",
        "网关": "central gateway 域控制器",
        "OTA": "远程升级 空中下载 远程更新",
        "V2X": "车联网 车路协同",
        "GPS": "定位 导航 位置",
        "蓝牙": "Bluetooth 无线连接 短距离通信",
        "WiFi": "无线局域网 无线网络",
        "4G": "蜂窝网络 移动通信 LTE",
        "5G": "蜂窝网络 移动通信",
        "USB": "通用串行总线 外接接口",
        "安全": "security 安保 防护",
        "攻击": "入侵 渗透 破解 绕过",
        "漏洞": "弱点 脆弱性 CVE",
        "威胁": "风险 攻击 危害",
        "损害": "伤害 损失 危害 破坏",
        "风险": "攻击可能性 可行性 威胁等级",
    }

    @staticmethod
    def _expand_query(query: str) -> str:
        """对查询进行同义词扩展，提升语义召回。"""
        expanded = query
        for term, syns in BM25RAGKnowledgeBase.SYNONYM_MAP.items():
            if term in query:
                expanded += " " + syns
        return expanded

    # ── 文件解析 ────────────────────────────────────

    @staticmethod
    def _safe_read_text(path: str) -> str:
        for enc in ("utf-8", "utf-8-sig", "gbk", "gb18030"):
            try:
                with open(path, "r", encoding=enc) as f:
                    return f.read()
            except Exception:
                continue
        return ""

    @staticmethod
    def _load_json_text(path: str) -> str:
        try:
            with open(path, "r", encoding="utf-8") as f:
                obj = json.load(f)
            if isinstance(obj, (dict, list)):
                return json.dumps(obj, ensure_ascii=False)
            return str(obj)
        except Exception:
            return BM25RAGKnowledgeBase._safe_read_text(path)

    @staticmethod
    def _parse_jsonl(path: str) -> list[dict]:
        """解析 JSONL 预切块文件，每条记录返回一个 metadata dict。"""
        chunks: list[dict] = []
        try:
            with open(path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        rec = json.loads(line)
                    except json.JSONDecodeError:
                        continue
                    text = rec.get("text", "")
                    if not text or len(text.strip()) < 20:
                        continue
                    chunks.append({
                        "text": text.strip(),
                        "library_name": rec.get("library_name", ""),
                        "source_name": rec.get("source_name", ""),
                        "source_type": rec.get("source_type", ""),
                        "tara_phase": rec.get("tara_phase", []),
                        "agent_scope": rec.get("agent_scope", []),
                        "knowledge_role": rec.get("knowledge_role", ""),
                    })
        except Exception as e:
            logger.warning("JSONL 解析失败 %s: %s", path, e)
        return chunks

    @staticmethod
    def _parse_pdf(path: str) -> str:
        try:
            import fitz
        except ImportError:
            logger.warning("pymupdf 未安装，跳过 PDF: %s", path)
            return ""
        try:
            doc = fitz.open(path)
            pages = []
            for page in doc:
                t = page.get_text()
                if t:
                    pages.append(t)
            doc.close()
            return "\n\n".join(pages)
        except Exception as e:
            logger.warning("PDF 解析失败 %s: %s", path, e)
            return ""

    @staticmethod
    def _parse_excel(path: str) -> str:
        try:
            import openpyxl
        except ImportError:
            logger.warning("openpyxl 未安装，跳过 Excel: %s", path)
            return ""
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            lines = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    line = " | ".join(str(c) for c in row if c is not None)
                    if line.strip():
                        lines.append(line)
            wb.close()
            return "\n".join(lines)
        except Exception as e:
            logger.warning("Excel 解析失败 %s: %s", path, e)
            return ""

    def _read_file(self, path: str) -> str:
        ext = os.path.splitext(path)[1].lower()
        if ext == ".pdf":
            return self._parse_pdf(path)
        elif ext == ".xlsx":
            return self._parse_excel(path)
        elif ext in {".txt", ".md"}:
            return self._safe_read_text(path)
        elif ext == ".json":
            return self._load_json_text(path)
        return ""

    # ── 切块（仅用于非 JSONL 源文件）────────────────

    @staticmethod
    def _chunk_text(text: str, chunk_size: int = 500, overlap: int = 50) -> list[str]:
        """按段落边界切分，保留语义完整性。"""
        if not text:
            return []
        if len(text) < chunk_size:
            return [text.strip()]
        paragraphs = re.split(r"\n\s*\n", text)
        paragraphs = [p.strip() for p in paragraphs if p.strip()]
        chunks: list[str] = []
        buffer = ""
        for para in paragraphs:
            if len(para) > chunk_size:
                if buffer:
                    chunks.append(buffer)
                    buffer = ""
                start = 0
                while start < len(para):
                    end = min(start + chunk_size, len(para))
                    chunks.append(para[start:end])
                    start = end - overlap if end < len(para) else len(para)
                continue
            if not buffer:
                buffer = para
            elif len(buffer) + len(para) + 2 <= chunk_size:
                buffer += "\n\n" + para
            else:
                chunks.append(buffer)
                buffer = para
        if buffer:
            if len(buffer) < 50 and chunks:
                chunks[-1] += "\n\n" + buffer
            else:
                chunks.append(buffer)
        return chunks

    # ── BM25 实现（倒排索引加速）────────────────────

    class _BM25Okapi:
        """BM25 Okapi 评分器，使用倒排索引避免全表扫描。"""
        def __init__(self, corpus: list[list[str]], k1: float = 1.5, b: float = 0.75):
            self.k1 = k1
            self.b = b
            self.corpus = corpus
            self.corpus_size = len(corpus)
            self.doc_lens = [len(doc) for doc in corpus]
            self.avgdl = sum(self.doc_lens) / self.corpus_size if self.corpus_size else 1.0

            from collections import Counter, defaultdict
            df: Counter[str] = Counter()
            # 倒排索引：term → [(doc_idx, tf), ...]
            inv: dict[str, list[tuple[int, int]]] = defaultdict(list)

            for i, doc in enumerate(corpus):
                seen_terms: set[str] = set()
                # 统计当前文档的词频
                local_tf: dict[str, int] = {}
                for term in doc:
                    local_tf[term] = local_tf.get(term, 0) + 1
                for term, tf in local_tf.items():
                    inv[term].append((i, tf))
                    seen_terms.add(term)
                for term in seen_terms:
                    df[term] += 1

            self.idf: dict[str, float] = {
                term: math.log((self.corpus_size - n + 0.5) / (n + 0.5) + 1.0)
                for term, n in df.items()
            }
            self.inverted_index: dict[str, list[tuple[int, int]]] = dict(inv)

        def get_scores(self, query: list[str]) -> list[float]:
            scores = [0.0] * self.corpus_size
            for q in query:
                idf_q = self.idf.get(q)
                if idf_q is None or idf_q == 0.0:
                    continue
                postings = self.inverted_index.get(q)
                if not postings:
                    continue
                k1 = self.k1
                b = self.b
                avgdl = self.avgdl
                doc_lens = self.doc_lens
                for doc_idx, tf in postings:
                    scores[doc_idx] += idf_q * (tf * (k1 + 1)) / (tf + k1 * (1 - b + b * doc_lens[doc_idx] / avgdl))
            return scores

    # ── 加载 ─────────────────────────────────────────

    def load(self, force: bool = False) -> None:
        with self._lock:
            if self._loaded and not force:
                return

            import jieba
            records: list[dict] = []

            for category, d in self.rag_dirs.items():
                if not os.path.isdir(d):
                    continue
                for p in glob.glob(os.path.join(d, "**", "*"), recursive=True):
                    if not os.path.isfile(p):
                        continue
                    ext = os.path.splitext(p)[1].lower()

                    if ext == ".jsonl":
                        jsonl_recs = self._parse_jsonl(p)
                        for r in jsonl_recs:
                            r["category"] = category
                            r["path"] = p
                        records.extend(jsonl_recs)
                        if jsonl_recs:
                            logger.info(f"  JSONL {os.path.basename(p)}: {len(jsonl_recs)} 条")
                        continue

                    text = self._read_file(p)
                    if not text or len(text.strip()) < 20:
                        continue
                    chunks = self._chunk_text(text)
                    for ci, chunk in enumerate(chunks):
                        records.append({
                            "text": chunk,
                            "category": category,
                            "path": p,
                            "library_name": "",
                            "source_name": "",
                            "source_type": "",
                            "tara_phase": [],
                            "agent_scope": [],
                            "knowledge_role": "",
                        })

            if not records:
                self.records = []
                self._bm25 = None
                logger.warning("RAG: 未找到可索引的文档")
                return

            logger.info(f"BM25 RAG: jieba 分词 {len(records)} 个文档块 ...")
            tokenized = [list(jieba.cut(r["text"])) for r in records]

            self.records = records
            self._bm25 = self._BM25Okapi(tokenized, k1=self.k1, b=self.b)
            logger.info(f"BM25 RAG 索引完成: {len(records)} 个文档块, {len(self._bm25.idf)} 个词项")

            self._loaded = True

    # ── 检索 ────────────────────────────────────────

    def retrieve(self, query: str, top_k: int = 5, max_chars: int = 1800,
                 library_filter: str | list[str] | None = None) -> str:
        """同义扩展 → BM25 评分 → 阶段过滤 → top-K。

        Args:
            query: 查询文本
            top_k: 返回最相似的 K 个 chunk
            max_chars: 返回文本的最大字符数
            library_filter: 可选的 library_name 过滤。
                None=不过滤，"damage_cases"=只搜损害案例，
                ["threat_patterns", "vulnerabilities"]=搜多个库。
                空字符串的 chunk（非 JSONL 来源）始终保留。
        """
        if not ENABLE_RAG:
            return "[RAG已关闭]"
        if not self._loaded:
            self.load()
        if not self.records or self._bm25 is None:
            return "[RAG为空]"

        import jieba

        # 同义扩展
        expanded_query = self._expand_query(query)

        # BM25 评分
        query_tokens = list(jieba.cut(expanded_query))
        scores = self._bm25.get_scores(query_tokens)

        # library_filter 过滤（置零非目标库的分数）
        if library_filter is not None:
            if isinstance(library_filter, str):
                filter_set = {library_filter}
            else:
                filter_set = set(library_filter)
            for i, r in enumerate(self.records):
                lib = r.get("library_name", "")
                if lib not in filter_set and lib != "":
                    scores[i] = -1.0

        # 取 top-K
        top_indices = sorted(
            range(len(scores)), key=lambda i: scores[i], reverse=True
        )

        results: list[str] = []
        total = 0
        for idx in top_indices:
            if scores[idx] <= 0:
                break
            snippet = self.records[idx]["text"][:500].strip()
            src = self.records[idx].get("source_name", "") or self.records[idx].get("category", "")
            line = f"[来源:{src}] {snippet}"
            if total + len(line) > max_chars:
                break
            results.append(line)
            total += len(line)
            if len(results) >= top_k:
                break

        return "\n---\n".join(results) if results else "[RAG未命中]"


rag_kb = BM25RAGKnowledgeBase(RAG_DIRS, k1=RAG_BM25_K1, b=RAG_BM25_B)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 4) 数据结构
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@dataclass
class TaskUnit:
    task_id: str
    function: str
    asset_name: str
    asset_type: str
    attribute_name: str
    threat_type: str
    target: str = ""
    source: str = ""

    damage_scenario: str = ""
    threat_scenarios: str = ""
    attack_paths: list[str] = field(default_factory=list)
    selected_path_frameworks: list[dict] = field(default_factory=list)

    influence_parameters: dict = field(default_factory=dict)
    impact_value: int = 1
    impact_level: str = "Negligible"

    attack_details: list[dict] = field(default_factory=list)

    risk_value: int = 1
    risk_treatment: str = "接受"

    errors: list[str] = field(default_factory=list)


def _task_to_dict(t: TaskUnit) -> dict:
    return {
        "task_id": t.task_id,
        "function": t.function,
        "asset_name": t.asset_name,
        "asset_type": t.asset_type,
        "attribute_name": t.attribute_name,
        "threat_type": t.threat_type,
        "target": t.target,
        "source": t.source,
        "damage_scenario": t.damage_scenario,
        "threat_scenarios": t.threat_scenarios,
        "attack_paths": t.attack_paths,
        "selected_path_frameworks": t.selected_path_frameworks,
        "influence_parameters": t.influence_parameters,
        "impact_value": t.impact_value,
        "impact_level": t.impact_level,
        "attack_details": t.attack_details,
        "risk_value": t.risk_value,
        "risk_treatment": t.risk_treatment,
        "errors": t.errors,
    }


def _task_from_dict(d: dict) -> TaskUnit:
    return TaskUnit(
        task_id=d.get("task_id", ""),
        function=d.get("function", ""),
        asset_name=d.get("asset_name", ""),
        asset_type=d.get("asset_type", ""),
        attribute_name=d.get("attribute_name", ""),
        threat_type=d.get("threat_type", ""),
        target=d.get("target", ""),
        source=d.get("source", ""),
        damage_scenario=d.get("damage_scenario", ""),
        threat_scenarios=str(d.get("threat_scenarios", "")),
        attack_paths=list(d.get("attack_paths", [])),
        selected_path_frameworks=list(d.get("selected_path_frameworks", [])),
        influence_parameters=dict(d.get("influence_parameters", {})),
        impact_value=int(d.get("impact_value", 1)),
        impact_level=str(d.get("impact_level", "Negligible")),
        attack_details=list(d.get("attack_details", [])),
        risk_value=int(d.get("risk_value", 1)),
        risk_treatment=str(d.get("risk_treatment", "接受")),
        errors=list(d.get("errors", [])),
    )


def _tasks_to_dict(tasks: list[TaskUnit]) -> list[dict]:
    return [_task_to_dict(t) for t in tasks]


def _tasks_from_dict(data: list[dict]) -> list[TaskUnit]:
    return [_task_from_dict(d) for d in data]


@dataclass
class TopologyElementVO:
    id: str
    name: str = ""
    type: int | None = None
    color: str = ""
    source_id: str = ""
    target_id: str = ""
    is_gateway: bool = False
    ids: list[str] = field(default_factory=list)


@dataclass
class PathNodeVO:
    component_name_id: str
    component_name: str
    pre_component_name_id: str
    pre_component_name: str
    line_id: str
    color: str
    is_gateway: bool


class TopologyElementType:
    COMPONENT = 1
    GATEWAY = 2
    EXTERNAL_COMPONENT = 3
    LINE = 4


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 5) 通用工具
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def safe_parse_json(text: Any) -> Any:
    if text is None:
        raise ValueError("输入为空")
    if not isinstance(text, str):
        text = str(text)

    cleaned = text.strip()
    cleaned = re.sub(r"<think.*?</think.*?>", "", cleaned, flags=re.DOTALL).strip()

    md_match = re.search(r"```(?:json|JSON)?\s*\n?(.*?)```", cleaned, re.DOTALL)
    if md_match:
        cleaned = md_match.group(1).strip()

    try:
        return json.loads(cleaned)
    except Exception:
        pass

    for pattern in [r"(\[[\s\S]*\])", r"(\{[\s\S]*\})"]:
        m = re.search(pattern, cleaned)
        if m:
            try:
                return json.loads(m.group(1))
            except Exception:
                continue

    raise ValueError(f"JSON解析失败，原文前200字: {cleaned[:200]}")


def llm_call_with_semaphore(system_prompt: str, user_prompt: str, retries: int = 2) -> str:
    with llm_semaphore:
        for attempt in range(retries + 1):
            try:
                endpoint = API_BASE.rstrip("/") + "/chat/completions"
                payload = {
                    "model": MODEL_NAME,
                    "messages": [
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": user_prompt},
                    ],
                    "temperature": TEMPERATURE,
                    "max_tokens": MAX_TOKENS,
                    "extra_body": {"enable_thinking": False},
                }
                body = json.dumps(payload).encode("utf-8")
                req = urlrequest.Request(
                    endpoint,
                    data=body,
                    headers={
                        "Authorization": f"Bearer {API_KEY}",
                        "Content-Type": "application/json",
                    },
                    method="POST",
                )
                with urlrequest.urlopen(req, timeout=REQUEST_TIMEOUT) as resp:
                    data = json.loads(resp.read().decode("utf-8"))
                content = (
                    data.get("choices", [{}])[0]
                    .get("message", {})
                    .get("content", "")
                )
                return str(content) if content is not None else ""
            except (urlerror.URLError, TimeoutError) as e:
                logger.warning(f"LLM网络错误: {type(e).__name__}: {e}")
            except Exception as e:
                logger.warning(f"LLM调用异常: {type(e).__name__}: {e}")

            if attempt < retries:
                wait = 2 ** attempt
                time.sleep(wait)

    return "{}"


def cached_llm_json(
    *,
    bucket: str,
    cache_key: Any,
    system_prompt: str,
    user_prompt: str,
    default: Any,
) -> tuple[Any, bool]:
    cached = semantic_cache.get(bucket, cache_key)
    if cached is not None:
        return cached, True

    raw = llm_call_with_semaphore(system_prompt, user_prompt)
    try:
        parsed = safe_parse_json(raw)
    except Exception:
        parsed = copy.deepcopy(default)

    semantic_cache.set(bucket, cache_key, parsed)
    return parsed, False


def normalize_attack_vector(value: str) -> str:
    s = (value or "").strip().lower()
    # Network（远程/蜂窝/卫星）
    if any(k in s for k in ["network", "网络", "remote", "ota", "蜂窝", "4g", "5g", "lte", "ethernet", "以太网", "cloud", "卫星定位"]):
        return "network"
    # Adjacent（短距无线）
    if any(k in s for k in ["adjacent", "相邻", "bluetooth", "蓝牙", "ble", "wifi", "wlan", "星闪", "v2x", "射频信道"]):
        return "adjacent"
    # Local（物理端口/本地接入）
    if any(k in s for k in ["local", "本地", "usb", "tf卡", "sd卡", "ic卡", "充电口", "obd", "内部通讯", "nfc", "登录"]):
        return "local"
    # Physical（侵入式调试）
    if any(k in s for k in ["physical", "物理", "debug", "jtag", "内部调试"]):
        return "physical"
    return "local"


def calculate_impact_value(influence_parameters: dict) -> int:
    if not influence_parameters:
        return 1
    v = max(
        int(influence_parameters.get("Safety", 0)),
        int(influence_parameters.get("Finance", 0)),
        int(influence_parameters.get("Operation", 0)),
        int(influence_parameters.get("Privacy", 0)),
    )
    return max(1, min(4, v))


def calculate_impact_level(impact_value: int) -> str:
    return {1: "Negligible", 2: "Moderate", 3: "Major", 4: "Severe"}.get(impact_value, "Negligible")


def calculate_attack_potential_score(params: dict) -> int:
    return sum(int(params.get(k, 0)) for k in AP_ALLOWED_VALUES.keys())


def calculate_attack_potential_level(total_score: int) -> str:
    # 用户要求的映射
    if total_score <= 9:
        return "High"
    if total_score <= 13:
        return "Medium"
    if total_score <= 19:
        return "Low"
    return "Very Low"


def calculate_cvss_score(params: dict) -> float:
    v = CVSS_V.get(normalize_attack_vector(params.get("attack_vector", "local")), 0.55)
    c = CVSS_C.get((params.get("complexity", "high") or "high").strip().lower(), 0.44)
    p = CVSS_P.get((params.get("privileges", "none") or "none").strip().lower(), 0.85)
    u = CVSS_U.get((params.get("user_interaction", "none") or "none").strip().lower(), 0.85)
    score = 8.22 * v * c * p * u
    return round(score, 2)


def calculate_cvss_level(score: float) -> str:
    if 2.96 <= score <= 3.89:
        return "High"
    if 2.00 <= score <= 2.95:
        return "Medium"
    if 1.06 <= score <= 1.99:
        return "Low"
    if 0.12 <= score <= 1.05:
        return "Very Low"
    return "Very Low" if score < 2.96 else "High"


def calculate_attack_vector_level(vector: str) -> str:
    return ATTACK_VECTOR_TO_LEVEL.get(normalize_attack_vector(vector), "Low")


def calculate_risk_value(impact_value: int, feasibility_level: str) -> int:
    return RISK_MATRIX.get((impact_value, feasibility_level), 1)


def get_risk_treatment(risk_value: int) -> str:
    if risk_value >= 5:
        return "风险缓解"
    if risk_value == 4:
        return "风险缓解"
    if risk_value == 3:
        return "风险缓解"
    if risk_value == 2:
        return "风险接受"
    return "风险接受"


def save_json(data: Any, path: str) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 5.5) Checkpoint（中间结果检查点）
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

CHECKPOINT_DIR = os.getenv("TARA_CHECKPOINT_DIR", os.path.join(OUTPUT_DIR, ".checkpoints"))
_CHECKPOINT_META_FILE = os.path.join(CHECKPOINT_DIR, "_meta.json")
_CHECKPOINT_FILES = {
    "stage_A_gen": "tasks_after_stageA.json",
    "stage_B_impact": "tasks_after_stageB.json",
    "stage_C_attack": "tasks_after_stageC.json",
}


def _checkpoint_path(stage: str) -> str:
    return os.path.join(CHECKPOINT_DIR, _CHECKPOINT_FILES[stage])


def _checkpoint_meta(method: str, total_tasks: int) -> dict:
    """生成当前运行的元信息快照。"""
    return {
        "method": method,
        "total_tasks": total_tasks,
        "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
        "checkpoint_version": "v10.1",
    }


def save_checkpoint(tasks: list[TaskUnit], stage: str, method: str) -> str:
    """将 tasks 序列化保存到对应阶段的 checkpoint 文件。"""
    os.makedirs(CHECKPOINT_DIR, exist_ok=True)
    path = _checkpoint_path(stage)
    data = {
        "meta": _checkpoint_meta(method, len(tasks)),
        "tasks": _tasks_to_dict(tasks),
    }
    save_json(data, path)
    logger.info("[Checkpoint] %s 已保存 -> %s (%d 个任务)", stage, path, len(tasks))
    return path


def load_checkpoint(stage: str) -> list[TaskUnit] | None:
    """加载指定阶段的 checkpoint，不存在或损坏时返回 None。"""
    path = _checkpoint_path(stage)
    if not os.path.isfile(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        tasks_data = data.get("tasks", [])
        if not tasks_data:
            logger.warning("[Checkpoint] %s 文件为空", stage)
            return None
        tasks = _tasks_from_dict(tasks_data)
        logger.info("[Checkpoint] %s 已加载 <- %s (%d 个任务)", stage, path, len(tasks))
        return tasks
    except Exception as e:
        logger.warning("[Checkpoint] %s 加载失败: %s，将重新运行该阶段", stage, e)
        return None


def resume_from_checkpoint(
    stages: list[str],
    method: str,
) -> tuple[int, list[TaskUnit] | None]:
    """按优先级检查已有的 checkpoint，找到第一个有效的 checkpoint。

    Args:
        stages: 按优先级排列的阶段名列表（从最新到最旧）。
        method: 期望的可行性方法。

    Returns:
        (stage_index, tasks) — stage_index 是需要继续的阶段的索引（0=S的开始），
        tasks 为已有 checkpoint 的 task 列表（stage_index > 0 时非 None）。
    """
    for i, stage in enumerate(stages):
        tasks = load_checkpoint(stage)
        if tasks is not None:
            # 校验元信息（方法是否一致）
            meta_path = _checkpoint_path(stage)
            try:
                with open(meta_path, "r", encoding="utf-8") as f:
                    meta = json.load(f).get("meta", {})
                saved_method = meta.get("method", "")
                if saved_method and saved_method != method:
                    logger.warning(
                        "[Checkpoint] %s 的方法(%s)与当前(%s)不一致，忽略",
                        stage, saved_method, method,
                    )
                    continue
            except Exception:
                pass

            # stage 在 stages 列表中的索引就是还需要跑的阶段范围
            remain_idx = i + 1
            if remain_idx >= len(stages):
                logger.info("[Checkpoint] 所有阶段已完成，直接使用最终结果")
                return remain_idx, tasks
            logger.info(
                "[Checkpoint] 从 %s checkpoint 恢复，将运行剩余 %d 个阶段",
                stage, len(stages) - remain_idx,
            )
            return remain_idx, tasks

    return 0, None


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 6) 输入解析（DFD.py 逻辑）
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def parse_dfd_json(input_dir: str) -> list[dict]:
    type_mapping = {
        "tm.Flow": "信号",
        "tm.Process": "部件",
        "tm.Store": "数据",
        "tm.Actor": "接口",
    }

    all_results: list[dict] = []
    json_files = glob.glob(os.path.join(input_dir, "*.json"))

    for file_path in json_files:
        with open(file_path, "r", encoding="utf-8") as f:
            try:
                data = json.load(f)
            except Exception:
                logger.warning(f"DFD解析失败，跳过: {os.path.basename(file_path)}")
                continue

        current_result = {"function": "", "assets": []}

        if "detail" in data and "diagrams" in data["detail"]:
            for diagram in data["detail"]["diagrams"]:
                if not current_result["function"] and "title" in diagram:
                    current_result["function"] = diagram["title"]

                if "diagramJson" not in diagram or "cells" not in diagram["diagramJson"]:
                    continue

                cells = diagram["diagramJson"]["cells"]
                id_to_cell = {c.get("id"): c for c in cells if c.get("id")}

                for cell in cells:
                    if cell.get("outOfScope") is True:
                        continue

                    raw_type = cell.get("type", "")
                    if raw_type == "tm.Boundary":
                        continue

                    asset_details = ""
                    finetermval_value = ""
                    for key, value in cell.items():
                        if key.startswith("propertyList") and isinstance(value, dict):
                            asset_details = value.get("assetDetails", "")
                            finetermval_value = value.get("finetermval", "")
                            break

                    if not finetermval_value:
                        continue

                    mapped_type = type_mapping.get(raw_type, raw_type)
                    target_value = ""
                    source_value = ""

                    if mapped_type == "部件":
                        target_value = finetermval_value

                    elif mapped_type == "数据":
                        cell_id = cell.get("id", "")
                        for flow_cell in cells:
                            if flow_cell.get("type") == "tm.Flow" and flow_cell.get("source", {}).get("id") == cell_id:
                                target_id = flow_cell.get("target", {}).get("id", "")
                                if target_id in id_to_cell:
                                    target_cell = id_to_cell[target_id]
                                    for k, v in target_cell.items():
                                        if k.startswith("propertyList") and isinstance(v, dict):
                                            target_value = v.get("finetermval", "")
                                            if target_value:
                                                break
                                if target_value:
                                    break

                    elif mapped_type == "信号":
                        target_id = cell.get("target", {}).get("id", "")
                        source_id = cell.get("source", {}).get("id", "")

                        if target_id in id_to_cell:
                            for k, v in id_to_cell[target_id].items():
                                if k.startswith("propertyList") and isinstance(v, dict):
                                    target_value = v.get("finetermval", "")
                                    if target_value:
                                        break

                        if source_id in id_to_cell:
                            for k, v in id_to_cell[source_id].items():
                                if k.startswith("propertyList") and isinstance(v, dict):
                                    source_value = v.get("finetermval", "")
                                    if source_value:
                                        break

                    item = {
                        "asset_type": mapped_type,
                        "asset_name": finetermval_value,
                        "assetDetails": asset_details,
                    }
                    if target_value:
                        item["target"] = target_value
                    if source_value:
                        item["source"] = source_value

                    current_result["assets"].append(item)

        if not current_result["function"]:
            current_result["function"] = os.path.splitext(os.path.basename(file_path))[0]

        if current_result["assets"]:
            all_results.append(current_result)

    logger.info(f"DFD解析完成: {len(all_results)} 个功能")
    return all_results


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 7) 拓扑解析
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class TopologyMapUtil:
    @staticmethod
    def list_element_vo(cells: list[dict]) -> list[TopologyElementVO]:
        result: list[TopologyElementVO] = []
        for cell in cells:
            ele = TopologyElementVO(id=cell.get("id"))
            shape = cell.get("shape", "")
            attrs = cell.get("attrs", {})

            if shape == "custom-rounded-rect":
                stroke = attrs.get("body", {}).get("stroke", "")
                if stroke == "#016AFF":
                    ele.type = TopologyElementType.GATEWAY
                    ele.is_gateway = True
                else:
                    ele.type = TopologyElementType.COMPONENT
                ele.color = stroke
                ele.name = attrs.get("text", {}).get("text", "")

            elif shape == "custom-rounded-rect-dash":
                ele.type = TopologyElementType.EXTERNAL_COMPONENT
                ele.color = attrs.get("body", {}).get("stroke", "")
                ele.name = attrs.get("text", {}).get("text", "")

            elif shape == "edge":
                ele.type = TopologyElementType.LINE
                ele.color = attrs.get("line", {}).get("stroke", "")
                ele.source_id = cell.get("source", {}).get("cell")
                ele.target_id = cell.get("target", {}).get("cell")

            result.append(ele)
        return result

    @staticmethod
    def get_distinct_components(element_vos: list[TopologyElementVO]) -> list[TopologyElementVO]:
        components = [
            e for e in element_vos
            if e.type in [
                TopologyElementType.COMPONENT,
                TopologyElementType.GATEWAY,
                TopologyElementType.EXTERNAL_COMPONENT,
            ] and e.name
        ]

        name_map: dict[str, list[TopologyElementVO]] = defaultdict(list)
        for c in components:
            name_map[c.name].append(c)

        distinct: list[TopologyElementVO] = []
        for _, elements in name_map.items():
            first = elements[0]
            first.ids = [e.id for e in elements]
            distinct.append(first)
        return distinct

    @staticmethod
    def generate_full_path(
        external_interfaces: list[dict],
        element_vos: list[TopologyElementVO],
        protocol_legends: list[dict],
        distinct_components: list[TopologyElementVO],
    ) -> list[list[PathNodeVO]]:
        many_to_one: dict[str, str] = {}
        for comp in distinct_components:
            for old_id in comp.ids:
                many_to_one[old_id] = comp.id

        valid_colors = [p.get("color") for p in protocol_legends if p.get("color")]
        lines = [e for e in element_vos if e.type == TopologyElementType.LINE and e.color in valid_colors]

        for line in lines:
            line.source_id = many_to_one.get(line.source_id, line.source_id)
            line.target_id = many_to_one.get(line.target_id, line.target_id)

        ext_names = [ext["related_component"] for ext in external_interfaces]
        ext_ids = [c.id for c in distinct_components if c.name in ext_names]

        com_to_lines: dict[str, set] = defaultdict(set)
        line_to_coms: dict[str, set] = defaultdict(set)

        for line in lines:
            if line.source_id:
                com_to_lines[line.source_id].add(line.color)
                line_to_coms[line.color].add(line.source_id)
            if line.target_id:
                com_to_lines[line.target_id].add(line.color)
                line_to_coms[line.color].add(line.target_id)

        starts = {k: v for k, v in com_to_lines.items() if k in ext_ids}
        com_dic = {c.id: c for c in distinct_components}

        return TopologyMapUtil.get_path_by_bfs(starts, com_to_lines, line_to_coms, com_dic)

    @staticmethod
    def get_path_by_bfs(
        starts: dict[str, set],
        com_to_lines: dict[str, set],
        line_to_coms: dict[str, set],
        com_dic: dict[str, TopologyElementVO],
    ) -> list[list[PathNodeVO]]:
        all_paths: list[list[PathNodeVO]] = []

        for start_id in starts.keys():
            queue = deque([(start_id, set(), [])])
            while queue:
                current_com_id, visited, path = queue.popleft()

                direct_points = set()
                line_ids = com_to_lines.get(current_com_id, set())
                for l_id in line_ids:
                    direct_points.update(line_to_coms.get(l_id, set()))

                for l_id in line_ids:
                    if l_id in visited:
                        continue

                    for next_com in line_to_coms.get(l_id, set()):
                        if next_com not in com_dic or next_com == current_com_id or next_com in visited:
                            continue

                        node = PathNodeVO(
                            component_name_id=next_com,
                            component_name=com_dic[next_com].name,
                            pre_component_name_id=current_com_id,
                            pre_component_name=com_dic[current_com_id].name,
                            line_id=l_id,
                            color=l_id,
                            is_gateway=com_dic[next_com].is_gateway,
                        )

                        new_path = copy.deepcopy(path)
                        new_path.append(node)
                        all_paths.append(new_path)

                        new_visited = set(visited)
                        new_visited.update(line_ids)
                        new_visited.update(direct_points)
                        queue.append((next_com, new_visited, new_path))

        return all_paths


def _determine_attack_vector(ext_interface: str, color_name_map: dict, path_nodes: list[PathNodeVO]) -> str:
    interface_lower = (ext_interface or "").lower()

    # Network
    if any(k in interface_lower for k in ["4g", "5g", "lte", "蜂窝", "ota", "cloud", "远程", "ethernet", "以太网", "卫星定位"]):
        return "network"
    # Adjacent
    if any(k in interface_lower for k in ["wifi", "wlan", "蓝牙", "bluetooth", "ble", "星闪", "v2x", "射频信道"]):
        return "adjacent"
    # Local
    if any(k in interface_lower for k in ["usb", "tf卡", "sd卡", "ic卡", "充电口", "obd", "内部通讯", "nfc", "local", "本地登录"]):
        return "local"
    # Physical
    if any(k in interface_lower for k in ["jtag", "debug", "内部调试", "uart", "spi", "i2c", "物理"]):
        return "physical"

    if path_nodes:
        protocol_name = color_name_map.get(path_nodes[0].color, "").lower()
        if any(k in protocol_name for k in ["wifi", "wlan", "4g", "5g", "ethernet", "以太网"]):
            return "network"
        if any(k in protocol_name for k in ["蓝牙", "ble", "nfc"]):
            return "adjacent"
        if any(k in protocol_name for k in ["usb", "obd"]):
            return "local"

    return "local"


def _norm_text(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).replace("\n", " ").replace("\r", " ").replace("\u3000", " ").strip()
    s = re.sub(r"\s+", " ", s)
    return s


def _is_checked(v: Any) -> bool:
    s = _norm_text(v).lower()
    if not s:
        return False
    if any(mark in s for mark in ["√", "✓", "☑", "✅"]):
        return True
    return s in {"1", "y", "yes", "true", "是", "有", "开", "通过"}


def _excel_col_to_idx(cell_ref: str) -> int:
    letters = "".join(ch for ch in cell_ref if ch.isalpha()).upper()
    idx = 0
    for ch in letters:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return max(0, idx - 1)


def _parse_xlsx_rows_stdlib(excel_path: str) -> list[dict]:
    """
    无 pandas/openpyxl 时，使用标准库解析 xlsx。
    返回：[{row_num:int, cells:{col_idx: value}}...]
    """
    ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    rel_ns = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}

    with zipfile.ZipFile(excel_path, "r") as zf:
        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in zf.namelist():
            root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in root.findall(".//x:si", ns):
                text = "".join((t.text or "") for t in si.findall(".//x:t", ns))
                shared_strings.append(text)

        # 找第一个工作表文件路径
        sheet_target = "worksheets/sheet1.xml"
        try:
            wb_root = ET.fromstring(zf.read("xl/workbook.xml"))
            sheets = wb_root.findall(".//x:sheets/x:sheet", ns)
            first_sheet = sheets[0] if sheets else None
            rid = first_sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id") if first_sheet is not None else None
            if rid and "xl/_rels/workbook.xml.rels" in zf.namelist():
                rel_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
                for rel in rel_root.findall(".//r:Relationship", rel_ns):
                    if rel.attrib.get("Id") == rid:
                        sheet_target = rel.attrib.get("Target", sheet_target)
                        break
        except Exception:
            pass

        sheet_path = "xl/" + sheet_target.lstrip("/")
        if sheet_path not in zf.namelist():
            # 兜底
            candidates = [n for n in zf.namelist() if n.startswith("xl/worksheets/sheet") and n.endswith(".xml")]
            if not candidates:
                return []
            sheet_path = sorted(candidates)[0]

        sheet_root = ET.fromstring(zf.read(sheet_path))
        out_rows: list[dict] = []

        for row in sheet_root.findall(".//x:sheetData/x:row", ns):
            row_num = int(row.attrib.get("r", "0") or 0)
            cells: dict[int, str] = {}
            for c in row.findall("x:c", ns):
                ref = c.attrib.get("r", "")
                col_idx = _excel_col_to_idx(ref) if ref else 0
                ctype = c.attrib.get("t", "")
                value = ""
                if ctype == "s":
                    v = c.find("x:v", ns)
                    if v is not None and v.text and v.text.isdigit():
                        ss_idx = int(v.text)
                        if 0 <= ss_idx < len(shared_strings):
                            value = shared_strings[ss_idx]
                elif ctype == "inlineStr":
                    is_node = c.find("x:is", ns)
                    if is_node is not None:
                        value = "".join((t.text or "") for t in is_node.findall(".//x:t", ns))
                else:
                    v = c.find("x:v", ns)
                    value = v.text if (v is not None and v.text is not None) else ""
                cells[col_idx] = _norm_text(value)
            out_rows.append({"row_num": row_num, "cells": cells})

        return out_rows


def _load_external_interfaces(excel_path: str) -> list[dict]:
    """读取外部接口清单：pandas -> openpyxl -> 标准库 xlsx 解析。"""
    # 1) pandas 路径
    if pd is not None:
        try:
            df = pd.read_excel(excel_path, header=1)
            df.columns = [_norm_text(c) for c in df.columns]

            direct_info_col = next((c for c in df.columns if "外部接口" in c and "信息" in c), None)
            direct_comp_col = next((c for c in df.columns if "关联部件" in c), None)
            if direct_info_col and direct_comp_col:
                direct_items = []
                for _, row in df.iterrows():
                    ext = _norm_text(row.get(direct_info_col, ""))
                    comp = _norm_text(row.get(direct_comp_col, ""))
                    if ext and comp and comp.lower() != "nan":
                        direct_items.append(
                            {
                                "external_interface": ext,
                                "related_component": comp,
                                "all_interfaces": [ext],
                            }
                        )
                if direct_items:
                    return direct_items

            comp_col = next((c for c in df.columns if "零部件" in c and "名称" in c), None)
            if comp_col is None and len(df.columns) >= 2:
                comp_col = df.columns[1]
            if comp_col is None:
                comp_col = df.columns[0] if len(df.columns) else None

            if comp_col is None:
                return []

            interface_columns = [c for c in df.columns if c != comp_col and _norm_text(c)]
            # 对齐原逻辑：倾向跳过前两个说明列
            if len(df.columns) > 2:
                interface_columns = [c for c in df.columns[2:] if c != comp_col and _norm_text(c)]

            external_interfaces: list[dict] = []
            for _, row in df.iterrows():
                comp = _norm_text(row.get(comp_col, ""))
                if not comp or comp.lower() == "nan":
                    continue
                active = [col for col in interface_columns if _is_checked(row.get(col, ""))]
                if active:
                    external_interfaces.append(
                        {
                            "external_interface": active[0],
                            "related_component": comp,
                            "all_interfaces": active,
                        }
                    )
            if external_interfaces:
                return external_interfaces
        except Exception as e:
            logger.warning(f"pandas读取外部接口失败，尝试其他方式: {e}")

    # 2) openpyxl 路径
    try:
        from openpyxl import load_workbook

        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active

        # 先尝试“外部接口信息 + 外部接口关联部件”直接映射格式
        direct_header_row = None
        info_idx = None
        comp_idx_direct = None
        for r in range(1, min(ws.max_row, 10) + 1):
            vals = [_norm_text(c.value) for c in ws[r]]
            for i, v in enumerate(vals):
                if "外部接口" in v and "信息" in v:
                    info_idx = i
                if "关联部件" in v:
                    comp_idx_direct = i
            if info_idx is not None and comp_idx_direct is not None:
                direct_header_row = r
                break

        if direct_header_row is not None:
            direct_items = []
            for row in ws.iter_rows(min_row=direct_header_row + 1, values_only=True):
                vals = [_norm_text(v) for v in row]
                ext = vals[info_idx] if info_idx < len(vals) else ""
                comp = vals[comp_idx_direct] if comp_idx_direct < len(vals) else ""
                if ext and comp:
                    direct_items.append(
                        {
                            "external_interface": ext,
                            "related_component": comp,
                            "all_interfaces": [ext],
                        }
                    )
            if direct_items:
                return direct_items

        # 优先寻找包含“零部件名称”的表头行
        header_row_idx = None
        comp_idx = None
        max_scan = min(ws.max_row, 20)
        for r in range(1, max_scan + 1):
            vals = [_norm_text(c.value) for c in ws[r]]
            for i, v in enumerate(vals):
                if "零部件" in v and "名称" in v:
                    header_row_idx = r
                    comp_idx = i
                    break
            if header_row_idx is not None:
                break

        if header_row_idx is None:
            header_row_idx = 2
            vals = [_norm_text(c.value) for c in ws[header_row_idx]]
            comp_idx = vals.index("零部件名称") if "零部件名称" in vals else 1

        header_vals = [_norm_text(c.value) for c in ws[header_row_idx]]
        interface_cols = [(i, h) for i, h in enumerate(header_vals) if i > comp_idx and h]

        external_interfaces: list[dict] = []
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            row_vals = [_norm_text(v) for v in row]
            if comp_idx >= len(row_vals):
                continue
            comp = row_vals[comp_idx]
            if not comp:
                continue
            active = [h for i, h in interface_cols if i < len(row_vals) and _is_checked(row_vals[i])]
            if active:
                external_interfaces.append(
                    {
                        "external_interface": active[0],
                        "related_component": comp,
                        "all_interfaces": active,
                    }
                )
        if external_interfaces:
            return external_interfaces
    except Exception:
        pass

    # 3) 标准库兜底（无 pandas / openpyxl）
    try:
        rows = _parse_xlsx_rows_stdlib(excel_path)
    except Exception as e:
        logger.warning(f"标准库解析Excel失败: {e}")
        return []

    if not rows:
        return []

    # A) “外部接口信息 + 外部接口关联部件”直接映射格式
    direct_header = None
    info_idx = None
    comp_idx_direct = None
    for r in rows[:20]:
        for idx, v in r["cells"].items():
            if "外部接口" in v and "信息" in v:
                info_idx = idx
            if "关联部件" in v:
                comp_idx_direct = idx
        if info_idx is not None and comp_idx_direct is not None:
            direct_header = r["row_num"]
            break

    if direct_header is not None:
        direct_items = []
        for r in rows:
            if r["row_num"] <= direct_header:
                continue
            cells = r["cells"]
            ext = _norm_text(cells.get(info_idx, ""))
            comp = _norm_text(cells.get(comp_idx_direct, ""))
            if ext and comp:
                direct_items.append(
                    {
                        "external_interface": ext,
                        "related_component": comp,
                        "all_interfaces": [ext],
                    }
                )
        if direct_items:
            return direct_items

    # B) 勾选矩阵格式
    header_row_num = None
    comp_idx = None
    for r in rows[:30]:
        for idx, v in r["cells"].items():
            if "零部件" in v and "名称" in v:
                header_row_num = r["row_num"]
                comp_idx = idx
                break
        if header_row_num is not None:
            break

    if header_row_num is None:
        # 尝试第2行兜底
        row2 = next((r for r in rows if r["row_num"] == 2), None)
        if row2:
            header_row_num = 2
            for idx, v in row2["cells"].items():
                if "零部件" in v and "名称" in v:
                    comp_idx = idx
                    break
            if comp_idx is None:
                comp_idx = 1
        else:
            return []

    header_row = next((r for r in rows if r["row_num"] == header_row_num), None)
    if not header_row:
        return []

    interface_cols = [(idx, name) for idx, name in sorted(header_row["cells"].items()) if idx > comp_idx and name]

    external_interfaces: list[dict] = []
    for r in rows:
        if r["row_num"] <= header_row_num:
            continue
        cells = r["cells"]
        comp = _norm_text(cells.get(comp_idx, ""))
        if not comp:
            continue
        active = [name for idx, name in interface_cols if _is_checked(cells.get(idx, ""))]
        if active:
            external_interfaces.append(
                {
                    "external_interface": active[0],
                    "related_component": comp,
                    "all_interfaces": active,
                }
            )

    if not external_interfaces:
        logger.warning("外部接口Excel已读取，但未识别到勾选项；请检查表头与勾选符号。")
    return external_interfaces


def parse_topology_and_generate_frameworks(
    topology_file: str,
    external_interface_excel: str,
    asset_results: list[dict] | None = None,
) -> list[dict]:
    with open(topology_file, "r", encoding="utf-8") as f:
        topo_data = json.load(f).get("data", {})

    cells = topo_data.get("cells", [])
    protocol_legends = topo_data.get("lineData", [])
    color_name_map = {p.get("color"): p.get("name") for p in protocol_legends}

    if not os.path.exists(external_interface_excel):
        logger.warning(f"外部接口文件不存在: {external_interface_excel}")
        return []
    external_interfaces = _load_external_interfaces(external_interface_excel)
    if not external_interfaces:
        logger.warning("外部接口清单为空，攻击路径框架将为空。")
        return []

    # 从DFD解析结果中提取所有 target 名称，用于后续框架预匹配
    dfd_targets: set[str] = set()
    if asset_results:
        for func_data in asset_results:
            for asset in func_data.get("assets", []):
                t = (asset.get("target") or "").strip()
                if t:
                    dfd_targets.add(t)

    element_vos = TopologyMapUtil.list_element_vo(cells)
    distinct_components = TopologyMapUtil.get_distinct_components(element_vos)
    path_node_vos = TopologyMapUtil.generate_full_path(
        external_interfaces, element_vos, protocol_legends, distinct_components
    )

    source_target_map: dict[str, list[list[PathNodeVO]]] = defaultdict(list)
    for p in path_node_vos:
        source_id = p[0].pre_component_name_id
        target_id = p[-1].component_name_id
        source_target_map[f"{source_id}_{target_id}"].append(p)

    comp_to_interfaces: dict[str, list[str]] = defaultdict(list)
    for ext in external_interfaces:
        comp_to_interfaces[ext["related_component"]].append(ext["external_interface"])

    frameworks: list[dict] = []

    for _, paths in source_target_map.items():
        min_len = min(len(p) for p in paths) if paths else 0
        min_paths = [p for p in paths if len(p) == min_len]

        has_gateway = any(any(n.is_gateway for n in p) for p in min_paths)
        filtered = [p for p in min_paths if any(n.is_gateway for n in p)] if has_gateway else min_paths

        for p in filtered:
            start_comp_name = p[0].pre_component_name
            ext_list = comp_to_interfaces.get(start_comp_name, ["未知接口"])

            for ext in ext_list:
                desc = [f"[{ext}] -> {start_comp_name}"]
                for node in p:
                    proto = color_name_map.get(node.color, "未知协议")
                    desc.append(f" -> [{proto}] -> {node.component_name}")
                path_desc = "".join(desc)

                path_nodes = [start_comp_name] + [n.component_name for n in p]
                last_node = path_nodes[-1] if path_nodes else ""
                second_to_last_node = path_nodes[-2] if len(path_nodes) >= 2 else ""
                attack_vector = _determine_attack_vector(ext, color_name_map, p)

                # 预计算此框架匹配哪些 DFD target（last_node == target）
                matched_targets = [t for t in dfd_targets if t == last_node]

                frameworks.append({
                    "path_id": hashlib.sha1(path_desc.encode("utf-8")).hexdigest()[:16],
                    "path_description": path_desc,
                    "path_nodes": path_nodes,
                    "last_node": last_node,
                    "second_to_last_node": second_to_last_node,
                    "start_component": start_comp_name,
                    "external_interface": ext,
                    "components": [n.component_name for n in p],
                    "attack_vector": attack_vector,
                    "matched_targets": matched_targets,
                })

    unique = []
    seen = set()
    for fw in frameworks:
        if fw["path_description"] in seen:
            continue
        seen.add(fw["path_description"])
        unique.append(fw)

    logger.info(f"拓扑攻击路径框架生成完成: {len(unique)} 条")
    return unique


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 8) Prompt（生成+内嵌自评）
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

DS_SYSTEM_PROMPT = """你是汽车攻击路径专家，精通 ISO/SAE 21434 中的TARA方法，熟悉GB 44495法规和UN R155 法规。
你正在执行 TARA 分析中的损害场景建模任务。
损害场景是”涉及车辆或车辆功能且影响道路使用者的不良后果”。
在输出前必须进行自检，仅输出通过自检的结果。

## 输入说明

以下信息将在 user_prompt 中提供：功能名称、资产名称、资产类别、安全属性、威胁类型，以及相关参考知识（RAG）。

## 生成规则

损害场景必须同时包含以下三个要素（缺一不可），且必须结合该资产的功能和安全属性进行描述：

   1. 功能如何因为资产的安全属性被破坏而产生不良后果
   2. 对道路使用者的损害说明（必须明确指出对车主、驾驶员、乘员、行人或其他道路使用者的具体损害）
   3. 必须明确提到被破坏的资产名称

## 参考示例：

- “存储在信息娱乐系统中的个人信息(客户个人偏好)失去机密性,在未经客户同意的情况下,披露客户个人信息。”
- “车辆夜间行驶时，前照灯控制功能因’前照灯请求信号’完整性被破坏而意外关闭，导致驾驶员在无照明条件下高速行驶，与静止障碍物发生正面碰撞的风险。”

【自检3条件】
1) 相关项功能与不良后果的因果关系
2) 对道路使用者的具体危害
3) 涉及的目标资产

仅输出JSON：
{"damage_scenario": "..."}
"""

TS_SYSTEM_PROMPT = """你是汽车攻击路径专家，精通 ISO/SAE 21434 中的TARA方法，熟悉GB 44495法规和UN R155 法规。
你正在执行 TARA 分析中的威胁场景建模任务。
威胁场景是”为实现损害场景，资产的信息安全属性遭到破坏的潜在原因”。
在输出前必须进行自检，仅输出通过自检的结果。

## 输入说明

以下信息将在 user_prompt 中提供：功能名称、资产名称、资产类型、安全属性、威胁类型、损害场景，以及相关参考知识（RAG）。

## 生成规则

1. 威胁场景必须描述：
   - 明确指出被攻击的具体资产名称
   - 该资产被破坏的安全属性
   - 该安全属性被破坏的原因/攻击意图
2. 损害场景与资产名称、攻击者、攻击方法、攻击工具及攻击面之间的依赖关系应能被威胁场景包含或与威胁场景关联

## 参考示例：

- “攻击者从制动ECU方面对CAN消息进行欺骗，导致CAN消息的完整性缺失,从而导致制动功能的完整性缺失。”
- “伪装信号导致发送至电源开关控制器的’灯光请求’信号的数据通信完整性丢失,可能造成前照灯意外关闭”

【自检3条件】
1) 目标资产
2) 被破坏的信息安全属性
3) 信息安全属性被破坏的原因

仅输出JSON：
{"threat_scenarios": "..."}
"""

DST_SYSTEM_PROMPT = """你是汽车攻击路径专家，精通 ISO/SAE 21434 中的TARA方法，熟悉GB 44495法规和UN R155 法规。
你正在执行 TARA 分析中的损害场景建模和威胁场景建模任务。
损害场景是"涉及车辆或车辆功能且影响道路使用者的不良后果"。
威胁场景是"为实现损害场景，资产的信息安全属性遭到破坏的潜在原因"。
在输出前必须进行自检，仅输出通过自检的结果。

## 输入说明
以下信息将在 user_prompt 中提供：功能名称、资产名称、资产类别、安全属性、威胁类型、损害场景参考知识（RAG）、威胁场景参考知识（RAG）。

## 损害场景生成规则
损害场景必须同时包含以下三个要素（缺一不可），且必须结合该资产的功能和安全属性进行描述：
  1. 功能如何因为资产的安全属性被破坏而产生不良后果
  2. 对道路使用者的损害说明（必须明确指出对车主、驾驶员、乘员、行人或其他道路使用者的具体损害）
  3. 必须明确提到被破坏的资产名称

## 威胁场景生成规则
1. 威胁场景必须描述：
   - 明确指出被攻击的具体资产名称
   - 该资产被破坏的安全属性
   - 该安全属性被破坏的原因/攻击意图
2. 损害场景与资产名称、攻击者、攻击方法、攻击工具及攻击面之间的依赖关系应能被威胁场景包含或与威胁场景关联

## 参考示例：
- "存储在信息娱乐系统中的个人信息(客户个人偏好)失去机密性,在未经客户同意的情况下,披露客户个人信息。"
- "车辆夜间行驶时，前照灯控制功能因'前照灯请求信号'完整性被破坏而意外关闭，导致驾驶员在无照明条件下高速行驶，与静止障碍物发生正面碰撞的风险。"
- "攻击者从制动ECU方面对CAN消息进行欺骗，导致CAN消息的完整性缺失,从而导致制动功能的完整性缺失。"
- "伪装信号导致发送至电源开关控制器的'灯光请求'信号的数据通信完整性丢失,可能造成前照灯意外关闭"

【自检条件】
损害场景：1) 相关项功能与不良后果的因果关系 2) 对道路使用者的具体危害 3) 涉及的目标资产
威胁场景：1) 目标资产 2) 被破坏的信息安全属性 3) 信息安全属性被破坏的原因

仅输出JSON：
{"damage_scenario": "...", "threat_scenarios": "..."}
"""

AP_SYSTEM_PROMPT = """你是汽车攻击路径专家，精通 ISO/SAE 21434 中的TARA方法，熟悉GB 44495法规和UN R155 法规。
你正在执行 TARA 分析中的攻击路径建模任务。
攻击路径是”为实现威胁场景的一组蓄意活动”。必须采用攻击树分析，从外部攻击面开始，沿车辆拓扑逻辑连贯地推进到目标资产，最终实现给定的威胁场景。
在输出前必须进行自检，仅输出通过自检的结果。

## 输入说明

以下信息将在 user_prompt 中提供：功能名称、资产名称、资产类型、安全属性、威胁场景列表、攻击路径框架（path_frameworks），以及相关参考知识（RAG）。

## 生成规则

1. 每条攻击路径必须是完整步骤链，从外部攻击入口到目标资产，每条路径必须能直接实现给定的威胁场景
2. 必须引用真实攻击手段（如CAN报文注入、固件逆向、中间人攻击、重放攻击、伪造身份、DoS泛洪、OBD物理注入、蓝牙/蜂窝/USB/OTA入侵等）
3. 同一条威胁场景的多条攻击路径不要重复，避免冗余
4. 必须基于提供的路径框架（path_frameworks），路径框架中第一个节点都是外部攻击入口，先分析攻击路径框架是否符合现实逻辑，不符合的不用于生成，再将具体的攻击手段等填入框架中，构成完整的攻击链
5. 攻击路径长度限制在5句以内，攻击路径的最后一句是具体的攻击行为
6. 每条路径的技术手段必须针对资产类型的特点设计，对于部件类资产，攻击路径的起点是外部接口，终点是被攻击的ECU；对于数据类资产，攻击路径的起点是外部接口，终点是存储数据的ECU；对于信号类资产，攻击路径的起点是外部接口，终点是接收信号的ECU

## 参考示例：

- “1.利用蜂窝接口损害远程通信ECU;2.利用远程通信ECU的CAN通信损害网关ECU;3.网关ECU转发恶意制动请求信号。”
- “1.攻击者通过蜂窝网络接口入侵了导航ECU;2.被入侵的导航ECU发送恶意控制信号;3.网关ECU转发恶意控制信号至电源开关执行器;4.恶意信号伪装成灯光请求(关灯)。”
- “1.攻击者可以本地访问OBD连接器;2.攻击者通过OBD连接器发送恶意控制信号;3.网关ECU转发恶意信号至电源开关执行器;4.恶意信号伪装成灯光请求(关灯)。”

【自检3条件】
1) 攻击路径必须包含：起点（具体攻击面：外部接口）、逻辑步骤、终点（目标资产）
2) 必须基于提供的路径框架，节点一致且步骤链逻辑闭合。
3) 每条路径都要可执行、具体。

仅输出JSON：
{"attack_path": ["步骤链1", "步骤链2"]}
"""

IMPACT_SYSTEM_PROMPT = """你是汽车网络安全风险评估专家，精通 ISO/SAE 21434 影响评级方法
影响是因损害场景造成的损害程度或物理伤害程度的估计，请根据损害场景对Safety、Finance、Operation、Privacy 四个维度进行影响评级。
## 评级标准（取值 1～4 整数，必须逐字对照、严格遵守，不得主观臆断）
### Safety:
- 1 =没有受伤
- 2 =轻度、中度伤害
- 3 =严重的和有生命危险的伤害（可能生存）
- 4 =威胁生命的伤害（不确定是否幸存），致命的伤害
### Finance:
- 1 =经济损失导致的影响不大，后果可忽略不计，或与道路使用者无关
- 2 =经济损失导致不便的后果，受影响的道路使用者将能用有限的资源来克服
- 3 =导致经济上的大量损失，受影响的道路使用者将能够克服这些后果
- 4 =经济损失导致的灾难性后果，受影响的道路使用者可能无法克服
利益相关者：车主、驾驶员、乘员、行人、供应商、主机厂
### Operation:
- 1 =操作上的损坏导致车辆功能没有损害或无法感知的损害
- 2 =操作上的损坏导致了车辆功能的部分退化（例：用户满意度受到负面影响）
- 3 =操作上的损坏导致了车辆重要功能的丧失或受损（例：司机的重大烦扰）
- 4 =操作上的损坏导致了车辆核心功能的丧失或受损（例：车辆不工作或出现核心功能的意外行为，如启用跛行回家模式或
自主驾驶到一个非预期的位置）
### Privacy:
- 1 =隐私侵犯不会给道路使用者带来不便 a)泄露的信息不敏感并且很难识别到PII主体
- 2 =隐私侵犯给道路使用者带来很多不便 a)泄露的信息敏感但很难识别到PII主体；b)泄露的信息不敏感但很容易识别到
PII主体
- 3 =隐私侵犯给道路使用者带来很严重的影响 a)泄露的信息及其敏感但很难识别到PII主体；b)泄露的信息敏感而且很容
易识别到PII主体
- 4 =隐私侵犯会对道路使用者造成重大甚至不可逆转的影响。泄露的信息高度敏感，并且很容易识别到PII主体
考虑车主、驾驶员、乘员、行人、供应商、主机厂
## 强约束条件
- 每个维度必须严格对照评级标准打分
- 分值必须是整数 1~4
仅输出JSON：
{"influence_parameters": {"Safety":0,"Finance":0,"Operation":0,"Privacy":0}}
"""

AP_FEASIBILITY_SYSTEM_PROMPT = """你是ISO/SAE 21434攻击潜力评估专家。
请输出五维分值，且必须从给定离散值中选择：

## 评级标准（必须逐字对照、严格遵守，不得主观臆断）
### Exposure_time:
- 0=实现攻击行为的时间小于等于1天
- 1=实现攻击行为的时间小于等于1周
- 4=实现攻击行为的时间小于等于1个月
- 17=实现攻击行为的时间小于等于6个月
- 19=实现攻击行为的时间大于6个月
### Professional_experience:
- 0=外行：与专家或专业人士相比缺乏知识，没有特别的专长。例1：普通人使用公开的攻击逐步描述
- 3=熟悉产品或系统类型的安全行为。例2：有经验的业主，普通技术人员知道简单和流行的攻击
- 6=熟悉底层算法、协议、硬件、结构、安全行为、密码学、经典攻击等。例3：有经验的技术人员或工程师
- 9=一个攻击的不同步骤需要专家级别的不同专业知识。例4：多名经验丰富的工程师
### Required_information:
- 0=公共信息（例如互联网上获得的信息）
- 3=受限制的信息（制造商和供应商共享的内部文档）
- 7=机密信息（例如软件源代码、防盗控制系统相关信息）
- 11=—严格保密的信息（只有少数人知道的特定客户校准或内存映射）
### Opportunity_window:
- 0=十分高——无限：通过公共/不受信任网络的高可用性，无任何时间限制（远程攻击、互联网/蜂窝接口）
- 1=高——容易：高可用性和有限访问时间（蓝牙配对、远程软件更新）
- 4=中——有限的物理和/或逻辑访问（进入未上锁车辆、车载诊断端口）
- 10=低—困难：对相关项或组件的不切实际的访问（破解IC、暴力破解密钥）
### Required_equipment:
- 0=标准设备（笔记本电脑、CAN适配器、普通工具）
- 4=专业设备（高档示波器、信号发生器、硬件调试设备）
- 7=定制设备（厂家限制的工具、电子显微镜）
- 9=多重定制设备，攻击不同步骤需要不同类型的定制设备

仅输出JSON：
{"attack_parameters": {"Exposure_time":0,"Professional_experience":0,"Required_information":0,"Opportunity_window":0,"Required_equipment":0}}
"""

CVSS_FEASIBILITY_SYSTEM_PROMPT = """你是CVSS可利用性评估专家。
请根据攻击路径分析并输出以下3个CVSS可利用性参数（攻击向量已由外部接口映射确定，无需输出）：

## CVSS 可利用性评级标准（必须逐字对照、严格遵守，不得主观臆断）
### complexity:
- 低 (Low): 0.77 = 不存在专门的访问条件或可减轻的情况。攻击者可以期望针对易受攻击的组件获得可重复的成功
- 高 (High): 0.44 = 成功的攻击取决于攻击者无法控制的情况。也就是说，成功的攻击不能随意完成，但需要攻击者在准备或执行攻击时投入一定的精力，才能预期成功的攻击

### user_interaction:
- 无 (None): 0.85 = 攻击可以在没有任何用户交互的情况下完成
- 需求 (Required): 0.62 = 成功的攻击需要用户交互

### privileges:
- 无 (None): 0.85 = 未经授权的攻击者
- 低 (Low): 0.62 = 需要用户级访问
- 高 (High): 0.27 = 需要管理员或系统级访问

## 强约束条件
- 每个参数必须严格对照以上标准取值

仅输出JSON：
{"attack_parameters": {"complexity":0,"privileges":0,"user_interaction":0}}
"""

CVSS_BATCH_SYSTEM_PROMPT = """你是CVSS可利用性评估专家。
请根据给定的多条攻击路径，逐条分析并输出每条路径的CVSS可利用性参数。
攻击向量已由外部接口映射确定，在输入中提供。

## CVSS 可利用性评级标准（必须逐字对照、严格遵守，不得主观臆断）
### complexity:
- 低 (Low): 0.77 = 不存在专门的访问条件或可减轻的情况
- 高 (High): 0.44 = 成功的攻击取决于攻击者无法控制的情况

### user_interaction:
- 无 (None): 0.85 = 攻击可以在没有任何用户交互的情况下完成
- 需求 (Required): 0.62 = 成功的攻击需要用户交互

### privileges:
- 无 (None): 0.85 = 未经授权的攻击者
- 低 (Low): 0.62 = 需要用户级访问
- 高 (High): 0.27 = 需要管理员或系统级访问

## 强约束条件
- 每个参数必须严格对照以上标准取值
- 输出列表的长度必须与输入的攻击路径数量一致

仅输出JSON：
{"attack_parameters_list": [
    {"complexity": "Low", "privileges": "None", "user_interaction": "None"},
    {"complexity": "High", "privileges": "Low", "user_interaction": "Required"}
]}
"""

AP_BATCH_SYSTEM_PROMPT = """你是ISO/SAE 21434攻击潜力评估专家。
请根据给定的多条攻击路径，逐条分析并输出每条路径的攻击潜力五维参数。
分值必须从给定离散值中选择。

## 评级标准（必须逐字对照、严格遵守，不得主观臆断）
### Exposure_time:
- 0=实现攻击行为的时间小于等于1天
- 1=实现攻击行为的时间小于等于1周
- 4=实现攻击行为的时间小于等于1个月
- 17=实现攻击行为的时间小于等于6个月
- 19=实现攻击行为的时间大于6个月
### Professional_experience:
- 0=外行：与专家或专业人士相比缺乏知识，没有特别的专长
- 3=熟悉产品或系统类型的安全行为
- 6=熟悉底层算法、协议、硬件、结构、安全行为、密码学、经典攻击等
- 9=一个攻击的不同步骤需要专家级别的不同专业知识
### Required_information:
- 0=公共信息（例如互联网上获得的信息）
- 3=受限制的信息（制造商和供应商共享的内部文档）
- 7=机密信息（例如软件源代码、防盗控制系统相关信息）
- 11=严格保密的信息（只有少数人知道的特定客户校准或内存映射）
### Opportunity_window:
- 0=十分高——无限：通过公共/不受信任网络的高可用性
- 1=高——容易：高可用性和有限访问时间
- 4=中——有限的物理和/或逻辑访问
- 10=低——困难：对相关项或组件的不切实际的访问
### Required_equipment:
- 0=标准设备（笔记本电脑、CAN适配器、普通工具）
- 4=专业设备（高档示波器、信号发生器、硬件调试设备）
- 7=定制设备（厂家限制的工具、电子显微镜）
- 9=多重定制设备，攻击不同步骤需要不同类型的定制设备

## 强约束条件
- 每个参数必须严格对照以上标准取值
- 输出列表的长度必须与输入的攻击路径数量一致

仅输出JSON：
{"attack_parameters_list": [
    {"Exposure_time": 0, "Professional_experience": 3, "Required_information": 0, "Opportunity_window": 1, "Required_equipment": 0}
]}
"""


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 9) 生成与评估核心函数
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def generate_damage_scenario(task: TaskUnit, rag_context: str = "") -> str:
    user_prompt = f"""输入：
- function: {task.function}
- asset_type: {task.asset_type}
- asset_name: {task.asset_name}
- attribute_name: {task.attribute_name}
- threat_type: {task.threat_type}

参考知识（RAG）:
{rag_context}

请输出 damage_scenario。"""

    cache_key = {
        "function": task.function,
        "asset_type": task.asset_type,
        "asset_name": task.asset_name,
        "attribute_name": task.attribute_name,
        "threat_type": task.threat_type,
        "rag_context_hash": hashlib.md5((rag_context or "").encode("utf-8")).hexdigest(),
    }

    _t_start = time.perf_counter()
    parsed, cache_hit = cached_llm_json(
        bucket="damage_scenario",
        cache_key=cache_key,
        system_prompt=DS_SYSTEM_PROMPT,
        user_prompt=user_prompt,
        default={"damage_scenario": ""},
    )
    _add_phase_time("damage", time.perf_counter() - _t_start)

    if cache_hit:
        logger.debug(f"[cache] damage_scenario task={task.task_id}")

    return str(parsed.get("damage_scenario", "")).strip()


def generate_threat_scenarios(task: TaskUnit, rag_context: str = "") -> str:
    user_prompt = f"""输入：
- function: {task.function}
- asset_type: {task.asset_type}
- asset_name: {task.asset_name}
- attribute_name: {task.attribute_name}
- threat_type: {task.threat_type}
- damage_scenario: {task.damage_scenario}

参考知识（RAG）:
{rag_context}

请输出 threat_scenarios。"""

    cache_key = {
        "function": task.function,
        "asset_type": task.asset_type,
        "asset_name": task.asset_name,
        "attribute_name": task.attribute_name,
        "threat_type": task.threat_type,
        "damage_scenario": task.damage_scenario,
        "rag_context_hash": hashlib.md5((rag_context or "").encode("utf-8")).hexdigest(),
    }

    _t_start = time.perf_counter()
    parsed, cache_hit = cached_llm_json(
        bucket="threat_scenarios",
        cache_key=cache_key,
        system_prompt=TS_SYSTEM_PROMPT,
        user_prompt=user_prompt,
        default={"threat_scenarios": ""},
    )
    _add_phase_time("threat", time.perf_counter() - _t_start)

    if cache_hit:
        logger.debug(f"[cache] threat_scenarios task={task.task_id}")

    result = parsed.get("threat_scenarios", parsed.get("threat_scenario", ""))
    if isinstance(result, list):
        result = result[0] if result else ""
    return str(result).strip()


def generate_damage_and_threat(task: TaskUnit, rag_ds: str = "", rag_ts: str = "") -> tuple[str, str]:
    """合并生成 damage_scenario + threat_scenarios，一次 LLM 调用。"""
    user_prompt = f"""输入：
- function: {task.function}
- asset_type: {task.asset_type}
- asset_name: {task.asset_name}
- attribute_name: {task.attribute_name}
- threat_type: {task.threat_type}

损害场景参考知识（RAG）:
{rag_ds}

威胁场景参考知识（RAG）:
{rag_ts}

请输出 damage_scenario 和 threat_scenarios。"""

    cache_key = {
        "function": task.function,
        "asset_type": task.asset_type,
        "asset_name": task.asset_name,
        "attribute_name": task.attribute_name,
        "threat_type": task.threat_type,
        "rag_ds_hash": hashlib.md5((rag_ds or "").encode("utf-8")).hexdigest(),
        "rag_ts_hash": hashlib.md5((rag_ts or "").encode("utf-8")).hexdigest(),
    }

    _t_start = time.perf_counter()
    parsed, cache_hit = cached_llm_json(
        bucket="damage_threat",
        cache_key=cache_key,
        system_prompt=DST_SYSTEM_PROMPT,
        user_prompt=user_prompt,
        default={"damage_scenario": "", "threat_scenarios": ""},
    )
    elapsed = time.perf_counter() - _t_start
    _add_phase_time("damage", elapsed / 2)
    _add_phase_time("threat", elapsed / 2)

    if cache_hit:
        logger.debug(f"[cache] damage+threat task={task.task_id}")

    ds = str(parsed.get("damage_scenario", "")).strip()
    ts = str(parsed.get("threat_scenarios", parsed.get("threat_scenario", ""))).strip()
    return ds, ts


def generate_attack_paths(task: TaskUnit, frameworks: list[dict], rag_context: str = "") -> list[str]:
    fw_brief = [
        {
            "path_id": fw["path_id"],
            "path_description": fw["path_description"],
            "last_node": fw["last_node"],
            "attack_vector": fw["attack_vector"],
            "path_nodes": fw["path_nodes"],
        }
        for fw in frameworks[:8]
    ]

    user_prompt = f"""输入：
- function: {task.function}
- asset_type: {task.asset_type}
- asset_name: {task.asset_name}
- attribute_name: {task.attribute_name}
- threat_scenarios: {task.threat_scenarios}
- path_frameworks: {json.dumps(fw_brief, ensure_ascii=False)}

参考知识（RAG）:
{rag_context}

请输出 attack_path。"""

    cache_key = {
        "function": task.function,
        "asset_type": task.asset_type,
        "asset_name": task.asset_name,
        "attribute_name": task.attribute_name,
        "threat_scenarios": task.threat_scenarios,
        "path_frameworks": fw_brief,
        "rag_context_hash": hashlib.md5((rag_context or "").encode("utf-8")).hexdigest(),
    }

    _t_start = time.perf_counter()
    parsed, cache_hit = cached_llm_json(
        bucket="attack_path",
        cache_key=cache_key,
        system_prompt=AP_SYSTEM_PROMPT,
        user_prompt=user_prompt,
        default={"attack_path": []},
    )
    _add_phase_time("attack_path", time.perf_counter() - _t_start)

    if cache_hit:
        logger.debug(f"[cache] attack_path task={task.task_id}")

    paths = parsed.get("attack_path", parsed.get("attack_paths", []))
    if isinstance(paths, str):
        return [paths]
    if isinstance(paths, list):
        return [str(p).strip() for p in paths if str(p).strip()]
    return []


def generate_influence_parameters(task: TaskUnit) -> dict:
    user_prompt = f"""输入：
- function: {task.function}
- asset_name: {task.asset_name}
- asset_type: {task.asset_type}
- attribute_name: {task.attribute_name}
- threat_type: {task.threat_type}
- damage_scenario: {task.damage_scenario}

请输出 influence_parameters。"""

    cache_key = {
        "asset_name": task.asset_name,
        "asset_type": task.asset_type,
        "attribute_name": task.attribute_name,
        "damage_scenario": task.damage_scenario,
    }

    _t_start = time.perf_counter()
    parsed, _ = cached_llm_json(
        bucket="influence_parameters",
        cache_key=cache_key,
        system_prompt=IMPACT_SYSTEM_PROMPT,
        user_prompt=user_prompt,
        default={"influence_parameters": {"Safety": 0, "Finance": 0, "Operation": 0, "Privacy": 0}},
    )
    _add_phase_time("impact", time.perf_counter() - _t_start)

    params = parsed.get("influence_parameters", {})
    return {
        "Safety": int(params.get("Safety", 0)),
        "Finance": int(params.get("Finance", 0)),
        "Operation": int(params.get("Operation", 0)),
        "Privacy": int(params.get("Privacy", 0)),
    }


def generate_attack_parameters_attack_potential(task: TaskUnit, attack_path: str) -> dict:
    user_prompt = f"""输入：
- function: {task.function}
- asset_name: {task.asset_name}
- attribute_name: {task.attribute_name}
- threat_scenarios: {task.threat_scenarios}
- attack_path: {attack_path}
请输出 attack_parameters。"""

    cache_key = {
        "asset_name": task.asset_name,
        "attribute_name": task.attribute_name,
        "attack_path": attack_path,
    }

    _t_start = time.perf_counter()
    parsed, _ = cached_llm_json(
        bucket="attack_parameters_ap",
        cache_key=cache_key,
        system_prompt=AP_FEASIBILITY_SYSTEM_PROMPT,
            user_prompt=user_prompt,
            default={"attack_parameters": {k: 0 for k in AP_ALLOWED_VALUES}},
        )
    _add_phase_time("feasibility", time.perf_counter() - _t_start)

    params = parsed.get("attack_parameters", {})
    fixed = {}
    for k, allowed in AP_ALLOWED_VALUES.items():
        v = int(params.get(k, 0))
        # 强制吸附到最近合法值
        fixed[k] = min(allowed, key=lambda x: abs(x - v))
    return fixed


def generate_attack_parameters_cvss(task: TaskUnit, attack_path: str, attack_vector: str = "local") -> dict:
    """生成CVSS参数。攻击向量从外部接口映射预先确定（不调用大模型）。"""
    user_prompt = f"""输入：
- function: {task.function}
- asset_name: {task.asset_name}
- threat_scenarios: {task.threat_scenarios}
- attack_path: {attack_path}
- 攻击向量（已从外部接口映射确定）: {attack_vector}

请分析上述攻击路径，输出剩余3个CVSS可利用性参数。"""

    cache_key = {
        "asset_name": task.asset_name,
        "attack_path": attack_path,
    }

    _t_start = time.perf_counter()
    parsed, _ = cached_llm_json(
        bucket="attack_parameters_cvss",
        cache_key=cache_key,
        system_prompt=CVSS_FEASIBILITY_SYSTEM_PROMPT,
            user_prompt=user_prompt,
            default={"attack_parameters": {"complexity": "high", "privileges": "none", "user_interaction": "none"}},
        )
    _add_phase_time("feasibility", time.perf_counter() - _t_start)

    p = parsed.get("attack_parameters", {})
    return {
        "attack_vector": attack_vector,
        "complexity": str(p.get("complexity", "high")).strip().lower(),
        "privileges": str(p.get("privileges", "none")).strip().lower(),
        "user_interaction": str(p.get("user_interaction", "none")).strip().lower(),
    }


def generate_attack_parameters_cvss_batch(task: TaskUnit, attack_paths: list[tuple[str, str]]) -> list[dict]:
    """批量评估多条攻击路径的CVSS参数，一次LLM调用。

    Args:
        task: TaskUnit
        attack_paths: [(attack_path_text, attack_vector), ...]

    Returns:
        [{"attack_vector": ..., "complexity": ..., "privileges": ..., "user_interaction": ...}, ...]
    """
    paths_block = "\n\n".join(
        f"[路径 {i}]\n攻击路径: {ap}\n攻击向量: {vec}"
        for i, (ap, vec) in enumerate(attack_paths)
    )

    user_prompt = f"""输入：
- function: {task.function}
- asset_name: {task.asset_name}
- threat_scenarios: {task.threat_scenarios}

多条攻击路径如下（攻击向量已由外部接口映射确定）：
{paths_block}

请为每条攻击路径分别分析并输出CVSS可利用性参数。"""

    cache_key = {
        "asset_name": task.asset_name,
        "attack_paths": [ap for ap, _ in attack_paths],
    }

    _t_start = time.perf_counter()
    parsed, _ = cached_llm_json(
        bucket="attack_parameters_cvss_batch",
        cache_key=cache_key,
        system_prompt=CVSS_BATCH_SYSTEM_PROMPT,
        user_prompt=user_prompt,
        default={"attack_parameters_list": []},
    )
    _add_phase_time("feasibility", time.perf_counter() - _t_start)

    params_list = parsed.get("attack_parameters_list", [])
    results = []
    for i, (ap, vec) in enumerate(attack_paths):
        p = params_list[i] if i < len(params_list) else {}
        results.append({
            "attack_vector": vec,
            "complexity": str(p.get("complexity", "high")).strip().lower(),
            "privileges": str(p.get("privileges", "none")).strip().lower(),
            "user_interaction": str(p.get("user_interaction", "none")).strip().lower(),
        })
    return results


def generate_attack_parameters_ap_batch(task: TaskUnit, attack_paths: list[str]) -> list[dict]:
    """批量评估多条攻击路径的 Attack Potential 参数，一次LLM调用。"""
    paths_block = "\n\n".join(
        f"[路径 {i}]\n{ap}"
        for i, ap in enumerate(attack_paths)
    )

    user_prompt = f"""输入：
- function: {task.function}
- asset_name: {task.asset_name}
- attribute_name: {task.attribute_name}
- threat_scenarios: {task.threat_scenarios}

多条攻击路径如下：
{paths_block}

请为每条攻击路径分别分析并输出攻击潜力五维参数。"""

    cache_key = {
        "asset_name": task.asset_name,
        "attack_paths": attack_paths,
    }

    _t_start = time.perf_counter()
    parsed, _ = cached_llm_json(
        bucket="attack_parameters_ap_batch",
        cache_key=cache_key,
        system_prompt=AP_BATCH_SYSTEM_PROMPT,
        user_prompt=user_prompt,
        default={"attack_parameters_list": []},
    )
    _add_phase_time("feasibility", time.perf_counter() - _t_start)

    params_list = parsed.get("attack_parameters_list", [])
    results = []
    for i in range(len(attack_paths)):
        p = params_list[i] if i < len(params_list) else {}
        fixed = {}
        for k, allowed in AP_ALLOWED_VALUES.items():
            v = int(p.get(k, 0))
            fixed[k] = min(allowed, key=lambda x: abs(x - v))
        results.append(fixed)
    return results


# ── 智能体 RAG 查询生成 ────────────────────────────


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 10) 任务构建 + 路径筛选
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def build_task_units(asset_results: list[dict]) -> list[TaskUnit]:
    tasks: list[TaskUnit] = []
    idx = 0
    for func_data in asset_results:
        function_name = func_data.get("function", "")
        for asset in func_data.get("assets", []):
            asset_type = asset.get("asset_type", "")
            asset_name = asset.get("asset_name", "")
            target = asset.get("target", "")
            source = asset.get("source", "")

            attrs = SECURITY_ATTRIBUTES_MAP.get(asset_type, ["完整性"])
            for attr in attrs:
                threat_type = ATTRIBUTE_TO_THREAT.get(attr, "未知")
                tasks.append(
                    TaskUnit(
                        task_id=f"{idx:05d}",
                        function=function_name,
                        asset_name=asset_name,
                        asset_type=asset_type,
                        attribute_name=attr,
                        threat_type=threat_type,
                        target=target,
                        source=source,
                    )
                )
                idx += 1

    logger.info(f"任务单元构建完成: {len(tasks)}")
    return tasks


def filter_path_frameworks(task: TaskUnit, all_frameworks: list[dict]) -> list[dict]:
    target = (task.target or "").strip()
    source = (task.source or "").strip()
    is_signal = task.asset_type == "信号"

    selected: list[dict] = []

    for fw in all_frameworks:
        # 优先使用框架预计算的 matched_targets（与 DFD 解析结果联动）
        matched = fw.get("matched_targets", [])
        target_ok = bool(matched) and target in matched
        if not target_ok:
            # 回退到动态匹配（无 DFD 数据时的兼容）
            last_node = (fw.get("last_node") or "").strip()
            target_ok = bool(target) and last_node == target

        if not target_ok:
            continue

        # 信号类额外检查：source 必须等于路径倒数第二个节点
        if is_signal and source:
            second_to_last = fw.get("second_to_last_node", "")
            if source != second_to_last:
                continue

        selected.append(fw)

    # 回退策略：如果严格筛选为空，放宽到 target 命中路径字符串
    if not selected and target:
        selected = [fw for fw in all_frameworks if target in fw.get("path_description", "")]

    # 仍为空时，再按资产名尝试
    if not selected and task.asset_name:
        selected = [fw for fw in all_frameworks if task.asset_name in fw.get("path_description", "")]

    if len(selected) > 8:
        selected = selected[:8]

    return selected


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 11) 并发流水线
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def process_task_generation(
    task: TaskUnit,
    all_frameworks: list[dict],
    rag_cache: Optional[dict[str, dict[str, str]]] = None,
) -> TaskUnit:
    """阶段A：每个资产独立并发（损害->威胁->攻击路径生成）"""
    rag_cache = rag_cache or {}
    dmg_key = f"{task.asset_type}|{task.attribute_name}"
    try:
        rag_ds = rag_cache.get("damage", {}).get(dmg_key, "")
        rag_ts = rag_cache.get("threat", {}).get(dmg_key, "")
        task.damage_scenario, task.threat_scenarios = generate_damage_and_threat(task, rag_ds, rag_ts)
        if not task.damage_scenario:
            task.errors.append("damage_scenario为空")
        if not task.threat_scenarios:
            task.errors.append("threat_scenarios为空")

        selected = filter_path_frameworks(task, all_frameworks)
        task.selected_path_frameworks = selected

        rag_ap = rag_cache.get("attack_path", {}).get(task.asset_name, "")
        task.attack_paths = generate_attack_paths(task, selected, rag_ap)
        if not task.attack_paths:
            task.errors.append("attack_path为空")

    except Exception as e:
        task.errors.append(f"生成阶段异常: {e}")

    return task


def process_task_gen_impact(
    task: TaskUnit,
    all_frameworks: list[dict],
    rag_cache: dict[str, dict[str, str]],
) -> TaskUnit:
    """单资产 A+B 阶段：损害场景→威胁场景→攻击路径(缓存)→影响评级（不含可行性评估）。"""
    dmg_key = f"{task.asset_type}|{task.attribute_name}"
    try:
        # ── Stage A: 场景生成（合并 damage+threat，一次 LLM 调用）────────
        rag_ds = rag_cache.get("damage", {}).get(dmg_key, "")
        rag_ts = rag_cache.get("threat", {}).get(dmg_key, "")
        task.damage_scenario, task.threat_scenarios = generate_damage_and_threat(task, rag_ds, rag_ts)
        if not task.damage_scenario:
            task.errors.append("damage_scenario为空")
        if not task.threat_scenarios:
            task.errors.append("threat_scenarios为空")

        selected = filter_path_frameworks(task, all_frameworks)
        task.selected_path_frameworks = selected

        # 攻击路径缓存：同一 asset_name 的路径相同，跨功能复用
        cached_paths = _get_attack_path_cache(task.asset_name)
        if cached_paths is not None:
            task.attack_paths = cached_paths
            logger.debug(f"[path_cache] task={task.task_id} 复用 asset={task.asset_name} 的攻击路径")
        else:
            rag_ap = rag_cache.get("attack_path", {}).get(task.asset_name, "")
            task.attack_paths = generate_attack_paths(task, selected, rag_ap)
            if task.attack_paths:
                _set_attack_path_cache(task.asset_name, task.attack_paths)

        if not task.attack_paths:
            task.errors.append("attack_path为空")

        # ── Stage B: 影响评级 ──────────────────────────
        params = generate_influence_parameters(task)
        task.influence_parameters = params
        task.impact_value = calculate_impact_value(params)
        task.impact_level = calculate_impact_level(task.impact_value)

    except Exception as e:
        task.errors.append(f"A+B阶段异常: {e}")

    return task


def process_in_batches(items: list[Any], batch_size: int, worker_fn, stage_name: str) -> list[Any]:
    """连续批处理：上一批结束后再提交下一批。"""
    outputs: list[Any] = [None] * len(items)
    total = len(items)
    if total == 0:
        return outputs

    for start in range(0, total, batch_size):
        end = min(start + batch_size, total)
        batch = items[start:end]
        logger.info(f"[{stage_name}] 连续批处理 {start + 1}-{end}/{total}")

        with ThreadPoolExecutor(max_workers=min(MAX_WORKERS, len(batch))) as executor:
            future_map = {executor.submit(worker_fn, i, item): i for i, item in enumerate(batch, start=start)}
            for future in as_completed(future_map):
                i = future_map[future]
                try:
                    outputs[i] = future.result()
                except Exception as e:
                    outputs[i] = e

    return outputs


def stage_b_influence_batch(tasks: list[TaskUnit]) -> None:
    """阶段B：influence_parameters 连续批处理。"""

    def worker(index: int, task: TaskUnit):
        params = generate_influence_parameters(task)
        impact_value = calculate_impact_value(params)
        impact_level = calculate_impact_level(impact_value)
        return {
            "index": index,
            "params": params,
            "impact_value": impact_value,
            "impact_level": impact_level,
        }

    results = process_in_batches(tasks, BATCH_SIZE_INFLUENCE, worker, "Impact")
    for item in results:
        if isinstance(item, dict):
            task = tasks[item["index"]]
            task.influence_parameters = item["params"]
            task.impact_value = item["impact_value"]
            task.impact_level = item["impact_level"]


def stage_c_attack_batch(tasks: list[TaskUnit], method: str) -> None:
    """阶段C：批量评估攻击可行性 + 风险矩阵。

    attack_vector 无需 LLM 调用，直接映射；
    cvss / attack_potential 每个 task 只调一次 LLM（合并所有 attack_path）。
    """

    # ── attack_vector: 无 LLM 调用 ──────────────────
    if method == "attack_vector":
        for task in tasks:
            details = []
            for ai, ap_text in enumerate(task.attack_paths or []):
                vector = "local"
                if task.selected_path_frameworks:
                    fw = task.selected_path_frameworks[ai % len(task.selected_path_frameworks)]
                    vector = fw.get("attack_vector", "local")
                norm = normalize_attack_vector(vector)
                params = {"attack_vector": norm}
                score = ATTACK_VECTOR_SCORE.get(norm, 0.55)
                level = calculate_attack_vector_level(vector)
                risk_value = calculate_risk_value(task.impact_value, level)
                details.append({
                    "attack_path": ap_text,
                    "attack_parameters": params,
                    "feasibility_score": score,
                    "feasibility_level": level,
                    "risk_value": risk_value,
                    "risk_treatment": get_risk_treatment(risk_value),
                })
            task.attack_details = details
            if details:
                task.risk_value = max(d["risk_value"] for d in details)
            else:
                task.risk_value = calculate_risk_value(task.impact_value, "Very Low")
                task.risk_treatment = get_risk_treatment(task.risk_value)
        return

    # ── CVSS / Attack Potential: 逐 task 合并路径，一次 LLM 调用 ──
    jobs: list[dict] = []
    for ti, task in enumerate(tasks):
        if not task.attack_paths:
            task.attack_details = []
            task.risk_value = calculate_risk_value(task.impact_value, "Very Low")
            task.risk_treatment = get_risk_treatment(task.risk_value)
            continue

        paths_with_meta = []
        for ai, ap_text in enumerate(task.attack_paths):
            vector = "local"
            if task.selected_path_frameworks:
                fw = task.selected_path_frameworks[ai % len(task.selected_path_frameworks)]
                vector = fw.get("attack_vector", "local")
            paths_with_meta.append({"text": ap_text, "vector": vector})

        jobs.append({
            "task_index": ti,
            "paths": paths_with_meta,
        })

    if not jobs:
        return

    def worker(index: int, job: dict):
        task = tasks[job["task_index"]]
        paths = job["paths"]

        if method == "cvss":
            path_tuples = [(p["text"], p["vector"]) for p in paths]
            all_params = generate_attack_parameters_cvss_batch(task, path_tuples)
        else:  # attack_potential
            path_texts = [p["text"] for p in paths]
            all_params = generate_attack_parameters_ap_batch(task, path_texts)

        details = []
        for ai, p in enumerate(paths):
            params = all_params[ai] if ai < len(all_params) else {}
            if method == "cvss":
                score = calculate_cvss_score(params)
                level = calculate_cvss_level(score)
            else:
                score = calculate_attack_potential_score(params)
                level = calculate_attack_potential_level(score)
            risk_value = calculate_risk_value(task.impact_value, level)
            details.append({
                "attack_path": p["text"],
                "attack_parameters": params,
                "feasibility_score": score,
                "feasibility_level": level,
                "risk_value": risk_value,
                "risk_treatment": get_risk_treatment(risk_value),
            })
        return {"task_index": job["task_index"], "details": details}

    results = process_in_batches(jobs, BATCH_SIZE_ATTACK, worker, "Feasibility")

    for r in results:
        if not isinstance(r, dict):
            continue
        task = tasks[r["task_index"]]
        task.attack_details = r["details"]
        if r["details"]:
            task.risk_value = max(d["risk_value"] for d in r["details"])
        else:
            task.risk_value = calculate_risk_value(task.impact_value, "Very Low")
        task.risk_treatment = get_risk_treatment(task.risk_value)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 12) 报告构建
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def build_tara_report(tasks: list[TaskUnit], feasibility_method: str) -> list[dict]:
    function_map: dict[str, dict] = {}

    for task in tasks:
        if task.function not in function_map:
            function_map[task.function] = {"function": task.function, "assets": {}}

        asset_key = (task.asset_name, task.asset_type, task.target, task.source)
        assets = function_map[task.function]["assets"]

        if asset_key not in assets:
            assets[asset_key] = {
                "asset_name": task.asset_name,
                "asset_type": task.asset_type,
                "security_attributes": [],
            }

        detail = {
            "attribute_name": task.attribute_name,
            "threat_type": task.threat_type,
            "damage_scenario": task.damage_scenario,
            "threat_scenarios": task.threat_scenarios,
            "influence_parameters": task.influence_parameters,
            "impact_value": task.impact_value,
            "impact_level": task.impact_level,
            "feasibility_method": feasibility_method,
            "attack": task.attack_details,
        }
        if task.errors:
            detail["errors"] = task.errors

        assets[asset_key]["security_attributes"].append(detail)

    report = []
    for func_data in function_map.values():
        report.append({
            "function": func_data["function"],
            "assets": list(func_data["assets"].values()),
        })

    return report


def build_error_report(tasks: list[TaskUnit]) -> list[dict]:
    rows = []
    for task in tasks:
        if not task.errors:
            continue
        rows.append({
            "task_id": task.task_id,
            "function": task.function,
            "asset_name": task.asset_name,
            "asset_type": task.asset_type,
            "attribute_name": task.attribute_name,
            "threat_type": task.threat_type,
            "errors": task.errors,
        })
    return rows


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 13) 主流程
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def run_tara(
    data_flow_dir: str,
    topology_file: str,
    external_interface_excel: str,
    output_path: str,
    feasibility_method: str = FEASIBILITY_METHOD,
    max_assets_for_test: Optional[int] = None,
    skip_checkpoint_resume: bool = False,
) -> tuple[list[dict], list[dict]]:
    method = (feasibility_method or "attack_potential").strip().lower()
    if method not in {"attack_potential", "attack_vector", "cvss"}:
        logger.warning(f"未知可行性方法: {method}，回退为 attack_potential")
        method = "attack_potential"

    _pipeline_start = time.perf_counter()

    logger.info("=" * 72)
    logger.info("TARA v10 开始")
    logger.info(f"可行性评估方法: {method}")
    logger.info("=" * 72)

    # ─── Step 1: DFD资产识别 ─────────────────────────
    asset_results = parse_dfd_json(data_flow_dir)

    # 测试模式：仅取前N个DFD资产
    if max_assets_for_test is not None and max_assets_for_test > 0:
        limited_results = []
        count = 0
        for func_data in asset_results:
            assets = func_data.get("assets", [])
            if count >= max_assets_for_test:
                break
            remain = max_assets_for_test - count
            if len(assets) > remain:
                copy_func = dict(func_data)
                copy_func["assets"] = assets[:remain]
                limited_results.append(copy_func)
                count += remain
                break
            else:
                limited_results.append(func_data)
                count += len(assets)
        asset_results = limited_results
        logger.info(f"测试模式启用：仅使用前 {max_assets_for_test} 个DFD资产，实际纳入 {count} 个")

    # ─── Step 2: 拓扑攻击路径框架 ────────────────────
    path_frameworks = []
    if os.path.exists(topology_file) and os.path.exists(external_interface_excel):
        path_frameworks = parse_topology_and_generate_frameworks(topology_file, external_interface_excel, asset_results)
    else:
        logger.warning("拓扑文件或接口清单不存在，攻击路径框架为空")

    # ─── Step 3: 扩展属性 + RAG ──────────────────────
    tasks = build_task_units(asset_results)

    # 构建去重 RAG 缓存：相同 (asset_type, attribute_name) 只检索一次
    # cache 结构: {"damage": {"信号|完整性": "上下文...", ...},
    #              "threat": {"信号|完整性": "上下文...", ...},
    #              "attack_path": {"T-Box": "上下文...", ...}}
    rag_cache: dict[str, dict[str, str]] = {"damage": {}, "threat": {}, "attack_path": {}}
    if ENABLE_RAG:
        _build_start = time.perf_counter()

        # damage / threat：按 (asset_type, attribute_name) 去重
        for phase, phase_filter in [("damage", "damage_cases"), ("threat", "threat_patterns")]:
            seen: set[tuple[str, str]] = set()
            for t in tasks:
                key = (t.asset_type, t.attribute_name)
                if key in seen:
                    continue
                seen.add(key)
                str_key = f"{key[0]}|{key[1]}"
                query = f"{key[0]} {key[1]} {t.threat_type} 汽车网络安全 损害"
                if phase == "threat":
                    query = f"{key[0]} {key[1]} {t.threat_type} 攻击模式 威胁"
                rag_cache[phase][str_key] = rag_kb.retrieve(
                    query, top_k=RAG_TOP_K, max_chars=RAG_MAX_CONTEXT_CHARS,
                    library_filter=phase_filter,
                )

        # attack_path：按 asset_name 去重
        seen_assets: set[str] = set()
        for t in tasks:
            if t.asset_name in seen_assets:
                continue
            seen_assets.add(t.asset_name)
            query = f"{t.asset_name} {t.asset_type} {t.function} 攻击路径 漏洞 渗透"
            rag_cache["attack_path"][t.asset_name] = rag_kb.retrieve(
                query, top_k=RAG_TOP_K, max_chars=RAG_MAX_CONTEXT_CHARS,
                library_filter=["threat_patterns", "vulnerabilities"],
            )

        _build_elapsed = time.perf_counter() - _build_start
        logger.info(
            "RAG缓存构建完成: damage=%d组 threat=%d组 attack_path=%d组 (%.2fs)",
            len(rag_cache["damage"]), len(rag_cache["threat"]),
            len(rag_cache["attack_path"]), _build_elapsed,
        )
    else:
        logger.info("RAG已关闭")

    # ─── Checkpoint 恢复 ──────────────────────────────
    # 全流水线模式下只有一个 checkpoint 点：stage_C_attack（最终完成态）
    checkpoint_stages = ["stage_C_attack"]
    current_task_data: list[TaskUnit] | None = None
    resume_from = 0

    if not skip_checkpoint_resume:
        resume_from, current_task_data = resume_from_checkpoint(checkpoint_stages, method)
    else:
        logger.info("Checkpoint恢复已跳过（skip_checkpoint_resume=True）")

    # ─── Step 4-6: A+B per-asset 并行 + C 批量评估 ──
    if resume_from <= 0:
        # ── Stage A+B: per-asset 并行 ──────────────────
        logger.info(f"Stage A+B：每个资产独立完成生成+影响评级，资产数={len(tasks)}")
        completed_tasks: list[TaskUnit] = []

        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            future_map = {
                executor.submit(process_task_gen_impact, t, path_frameworks, rag_cache): t
                for t in tasks
            }
            done = 0
            total = len(future_map)
            for future in as_completed(future_map):
                t = future_map[future]
                done += 1
                try:
                    completed_tasks.append(future.result())
                except Exception as e:
                    t.errors.append(f"A+B阶段异常: {e}")
                    completed_tasks.append(t)
        completed_tasks.sort(key=lambda x: x.task_id)

        # ── Stage C: 攻击可行性批量评估 ─────────────────
        total_aps = sum(len(t.attack_paths or []) for t in completed_tasks)
        logger.info(f"Stage C：批量评估攻击可行性，总攻击路径数={total_aps}")
        stage_c_attack_batch(completed_tasks, method)

        save_checkpoint(completed_tasks, "stage_C_attack", method)
        current_task_data = completed_tasks
    else:
        logger.info("全流水线跳过（从 checkpoint 恢复，全部已完成）")

    assert current_task_data is not None, "当前任务数据为空，无法继续"

    # ─── Step 7: 报告 ────────────────────────────────
    report = build_tara_report(current_task_data, method)
    err_report = build_error_report(current_task_data)

    save_json(report, output_path)
    save_json(err_report, output_path.replace(".json", "_errors.json"))

    total_elapsed = time.perf_counter() - _pipeline_start
    logger.info("=" * 72)
    logger.info(f"TARA v10 完成: 总任务={len(current_task_data)}, 异常任务={len(err_report)}")
    logger.info(f"损害场景总运行时间: {_phase_times['damage']:.2f} 秒")
    logger.info(f"威胁场景总运行时间: {_phase_times['threat']:.2f} 秒")
    logger.info(f"攻击路径环节总运行时间: {_phase_times['attack_path']:.2f} 秒")
    logger.info(f"影响评级总运行时间: {_phase_times['impact']:.2f} 秒")
    logger.info(f"攻击可行性评级总运行时间: {_phase_times['feasibility']:.2f} 秒")
    logger.info(f"项目总运行时间: {total_elapsed:.2f} 秒 ({total_elapsed / 60:.2f} 分钟)")
    logger.info(f"主报告: {output_path}")
    logger.info(f"异常报告: {output_path.replace('.json', '_errors.json')}")
    logger.info("=" * 72)

    return report, err_report


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 14) 入口
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

if __name__ == "__main__":
    topology_file = r"D:\Jupyter profile\汽车信息安全风险评估\data\input\topology\拓扑图数据导出2026_5_9 10_43_29.json"
    external_interface_excel = r"D:\Jupyter profile\汽车信息安全风险评估\data\input\information\对外接口清单.xlsx"
    data_flow_dir = r"D:\Jupyter profile\汽车信息安全风险评估\data\input\DFD"

    # 你可以在这里切换：attack_potential / attack_vector / cvss
    method = os.getenv("TARA_FEASIBILITY_METHOD", FEASIBILITY_METHOD)
    # 测试时仅使用前N个DFD资产（默认1，设为0或负数则关闭）
    test_max_assets = int(os.getenv("TARA_TEST_MAX_ASSETS", "1"))
    if test_max_assets <= 0:
        test_max_assets = None
    # Checkpoint恢复：设 TARA_SKIP_CHECKPOINT=1 强制重跑所有阶段
    skip_cp = os.getenv("TARA_SKIP_CHECKPOINT", "1").strip() in {"1", "true", "True"}

    run_tara(
        data_flow_dir=data_flow_dir,
        topology_file=topology_file,
        external_interface_excel=external_interface_excel,
        output_path=os.path.join(OUTPUT_DIR, "tara_report_v10.json"),
        feasibility_method=method,
        max_assets_for_test=test_max_assets,
        skip_checkpoint_resume=skip_cp,
    )