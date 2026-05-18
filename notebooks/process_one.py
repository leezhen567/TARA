import openpyxl
import json
import os

DATA_RAW_PATH = r"D:/Jupyter profile/汽车信息安全风险评估/data/raw"
DATA_PROCESSED_PATH = r"D:/Jupyter profile/汽车信息安全风险评估/data/processed"

output = []

excel_file = os.path.join(DATA_RAW_PATH, "MY26", "动力系统", "2.1高压电源管理系统MY26TARA.xlsx")
output.append(f"文件: {excel_file}")
output.append(f"存在: {os.path.exists(excel_file)}")

wb = openpyxl.load_workbook(excel_file, data_only=True)
output.append(f"Sheet: {wb.sheetnames}")
ws = wb[wb.sheetnames[-1]]
output.append(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

row = 7
count = 0
while ws[f'A{row}'].value:
    count += 1
    row += 1
    if count > 100:
        break

output.append(f"数据行数: {count}")
wb.close()

with open('debug_output.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(output))