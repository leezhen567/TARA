"""
MY26文件夹批量TARA数据处理脚本
处理MY26文件夹中所有子文件夹里的Excel文件
"""

import openpyxl
import json
import os
import sys

DATA_RAW_PATH = r"D:/Jupyter profile/汽车信息安全风险评估/data/raw"
DATA_PROCESSED_PATH = r"D:/Jupyter profile/汽车信息安全风险评估/data/processed"
INTERFACE_FOLDER = "外部接口信息"
TOPOLOGY_FOLDER = "拓扑图"

def process_tara_data(ws):
    col_mapping = {
        'C': '一级功能',
        'D': '二级功能',
        'F': '资产',
        'H': '资产类别',
        'I': '安全属性',
        'L': '损害场景',
        'Y': '威胁场景',
        'Z': '攻击路径',
        'M': '安全',
        'O': '财产',
        'Q': '操作',
        'S': '隐私',
        'AA': '暴露时间',
        'AC': '专业经验',
        'AE': '所需信息',
        'AG': '机会窗口',
        'AI': '所需设备',
    }

    results = []
    row = 7

    while True:
        if not ws[f'A{row}'].value:
            break

        item = {}
        for col, name in col_mapping.items():
            cell = ws[f'{col}{row}']
            value = cell.value
            if value and isinstance(value, str) and value.startswith('='):
                value = ""
            if value == 0:
                item[name] = "0"
            elif value:
                item[name] = value
            else:
                item[name] = ""

        results.append(item)
        row += 1

    return results

def process_interface_data(file_path):
    if not os.path.exists(file_path):
        return []

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    interfaces = []
    row = 2

    while True:
        info = ws[f'A{row}'].value
        part = ws[f'B{row}'].value

        if not info and not part:
            break

        interfaces.append({
            "外部接口信息": info if info else "",
            "外部接口关联部件": part if part else ""
        })
        row += 1

    wb.close()
    return interfaces

def process_topology_data(file_path):
    if not os.path.exists(file_path):
        return []

    topology = []
    with open(file_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if line:
                topology.append(line)
    return topology

def process_file(excel_file, subfolder_name):
    file_name = os.path.basename(excel_file).replace('.xlsx', '')
    print(f"[{file_name}] 开始处理...", flush=True)

    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb[wb.sheetnames[-1]]
        print(f"[{file_name}] 读取数据...", flush=True)
        tara_data = process_tara_data(ws)
        wb.close()
        print(f"[{file_name}] 数据条数: {len(tara_data)}", flush=True)

        if not tara_data:
            print(f"[{file_name}] 无数据")
            return 0

        output_name = f"MY26_{file_name}"

        interface_file = os.path.join(DATA_RAW_PATH, INTERFACE_FOLDER, f"{file_name}.xlsx")
        topology_file = os.path.join(DATA_PROCESSED_PATH, TOPOLOGY_FOLDER, f"{file_name}.txt")

        interface_data = process_interface_data(interface_file)
        topology_data = process_topology_data(topology_file)

        for item in tara_data:
            item["外部接口"] = interface_data
            item["拓扑图"] = topology_data

        output_dir = os.path.join(DATA_PROCESSED_PATH, output_name)
        os.makedirs(output_dir, exist_ok=True)

        output_file = os.path.join(output_dir, f"{output_name}.json")
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(tara_data, f, ensure_ascii=False, indent=2)

        print(f"[{file_name}] 完成: {len(tara_data)} 条")
        return len(tara_data)

    except Exception as e:
        print(f"[{file_name}] 错误: {e}")
        import traceback
        traceback.print_exc()
        return 0

def main():
    base_path = os.path.join(DATA_RAW_PATH, "MY26")
    total_files = 0
    total_data = 0

    file_list = []
    for root, dirs, files in os.walk(base_path):
        for f in files:
            if f.endswith('.xlsx') and not f.startswith('~$'):
                excel_file = os.path.join(root, f)
                subfolder = os.path.basename(root)
                file_list.append((excel_file, subfolder))

    print(f"找到 {len(file_list)} 个文件", flush=True)

    for i, (excel_file, subfolder) in enumerate(file_list):
        print(f"[{i+1}/{len(file_list)}] ", end="", flush=True)
        count = process_file(excel_file, subfolder)
        total_files += 1
        total_data += count

    print(f"\n总计: 处理 {total_files} 个文件, {total_data} 条数据")

if __name__ == "__main__":
    main()