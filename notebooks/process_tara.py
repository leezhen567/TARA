"""
TARA数据处理通用脚本
处理欧胜传统车TARA-V1.5.xlsx文件，提取数据并保存为JSON

可修改的位置：
1. FILE_NAME: 原始数据文件名（不含路径）
2. OUTPUT_NAME: 输出JSON文件名（不含路径）
3. DATA_RAW_PATH: 原始数据文件夹路径
4. DATA_PROCESSED_PATH: 处理后数据保存路径
"""

import openpyxl
import json
import os

# ==================== 可修改的配置 ====================
FILE_NAME = "T9TARA -V1.2.xlsx"
OUTPUT_NAME = "T9TARA -V1.2"

DATA_RAW_PATH = r"D:/Jupyter profile/汽车信息安全风险评估/data/raw"
DATA_PROCESSED_PATH = r"D:/Jupyter profile/汽车信息安全风险评估/data/processed"

INTERFACE_FOLDER = "外部接口信息"
TOPOLOGY_FOLDER = "拓扑图"
# =====================================================

def get_title_from_merged(ws, start_row, end_row, col):
    """从合并单元格中获取标题（从下往上找第一个非空值）"""
    for row in range(end_row, start_row - 1, -1):
        cell = ws[f'{col}{row}']
        if cell.value:
            return cell.value
    return ""

def process_tara_data(wb):
    """处理TARA主数据"""
    ws = wb[wb.sheetnames[-1]]

    col_mapping = {
        'D': '功能',
        'F': '资产',
        'H': '资产类别',
        'I': '安全属性',
        'L': '损害场景',
        'Y': '威胁场景',
        'Z': '攻击路径',
        'M': '安全',
        'O': '财产',
        'Q': '操作',
        'S': '隐私',
        'AB': '暴露时间',
        'AD': '专业经验',
        'AF': '所需信息',
        'AH': '机会窗口',
        'AJ': '所需设备',
    }

    results = []
    row = 7

    while True:
        if not ws[f'A{row}'].value:
            break

        item = {}
        for col, name in col_mapping.items():
            cell = ws[f'{col}{row}']
            value = cell.value
            if value and isinstance(value, str) and value.startswith('='):
                value = ""
            if value == 0:
                item[name] = "0"
            elif value:
                item[name] = value
            else:
                item[name] = ""

        results.append(item)
        row += 1

    return results

def process_interface_data(file_path):
    """处理外部接口数据"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    interfaces = []
    row = 2

    while True:
        info = ws[f'A{row}'].value
        part = ws[f'B{row}'].value

        if not info and not part:
            break

        interfaces.append({
            "外部接口信息": info if info else "",
            "外部接口关联部件": part if part else ""
        })
        row += 1

    wb.close()
    return interfaces

def process_topology_data(file_path):
    """处理拓扑图数据"""
    topology = []

    with open(file_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if line:
                topology.append(line)

    return topology

def main():
    tara_file = os.path.join(DATA_RAW_PATH, FILE_NAME)
    interface_file = os.path.join(DATA_RAW_PATH, INTERFACE_FOLDER, FILE_NAME)
    topology_file = os.path.join(DATA_PROCESSED_PATH, TOPOLOGY_FOLDER, FILE_NAME.replace('.xlsx', '.txt'))

    output_dir = os.path.join(DATA_PROCESSED_PATH, OUTPUT_NAME)
    os.makedirs(output_dir, exist_ok=True)

    print(f"正在处理: {FILE_NAME}")

    wb = openpyxl.load_workbook(tara_file, data_only=True)
    tara_data = process_tara_data(wb)
    wb.close()
    print(f"  - TARA数据: {len(tara_data)} 条")

    interface_data = []
    if os.path.exists(interface_file):
        interface_data = process_interface_data(interface_file)
        print(f"  - 外部接口: {len(interface_data)} 条")
    else:
        print(f"  - 外部接口文件不存在: {interface_file}")

    topology_data = []
    if os.path.exists(topology_file):
        topology_data = process_topology_data(topology_file)
        print(f"  - 拓扑图: {len(topology_data)} 条")
    else:
        print(f"  - 拓扑图文件不存在: {topology_file}")

    for item in tara_data:
        item["外部接口"] = interface_data
        item["拓扑图"] = topology_data

    output_file = os.path.join(output_dir, f"{OUTPUT_NAME}.json")
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(tara_data, f, ensure_ascii=False, indent=2)

    print(f"\n已保存到: {output_file}")
    print(f"共处理 {len(tara_data)} 条数据")

if __name__ == "__main__":
    main()