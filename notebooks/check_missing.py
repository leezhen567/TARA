import openpyxl
import os

DATA_RAW_PATH = r"D:/Jupyter profile/汽车信息安全风险评估/data/raw"

missing_files = [
    "车身系统/1.1外灯子系统MY26TARA.xlsx",
    "车身系统/1.2内灯子系统MY26TARA.xlsx",
    "车身系统/1.3低速报警子系统MY26TARA.xlsx",
    "车身系统/1.4进入及门锁子系统MY26TARA.xlsx",
    "车身系统/1.5车辆防盗系统MY26TARA.xlsx",
    "车身系统/1.6防盗认证子系统MY26TARA.xlsx",
    "车身系统/1.7雨刮与清洗子系统MY26TARA.xlsx",
    "车身系统/1.8座椅控制子系统MY26TARA.xlsx",
    "车身系统/1.9车窗子系统MY26TARA.xlsx",
    "车身系统/1.19低压电源模式MY26TARA.xlsx",
    "车身系统/1.20车载冰箱MY26TARA.xlsx",
    "驾驶系统/4.1智能驾驶控制系统MY26TARA.xlsx",
]

output = []
for f in missing_files:
    file_path = os.path.join(DATA_RAW_PATH, "MY26", f)
    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[wb.sheetnames[-1]]
        row = 7
        count = 0
        while ws[f'A{row}'].value and count < 10:
            count += 1
            row += 1
        output.append(f"{f}: {count} rows")
        wb.close()
    else:
        output.append(f"{f}: NOT FOUND")

with open('missing_check.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(output))