import openpyxl
import json
import os
import sys
import pathlib

DATA_RAW_PATH = pathlib.Path("D:/Jupyter profile/汽车信息安全风险评估/data/raw")
DATA_PROCESSED_PATH = pathlib.Path("D:/Jupyter profile/汽车信息安全风险评估/data/processed")
INTERFACE_FOLDER = "外部接口信息"
TOPOLOGY_FOLDER = "拓扑图"

output_lines = []

def log_msg(msg):
    output_lines.append(msg)
    print(msg)

def process_tara_data(ws):
    col_mapping = {
        'D': '功能',
        'F': '资产',
        'H': '资产类别',
        'I': '安全属性',
        'L': '损害场景',
        'Y': '威胁场景',
        'Z': '攻击路径',
        'M': '安全',
        'O': '财产',
        'Q': '操作',
        'S': '隐私',
        'AB': '暴露时间',
        'AD': '专业经验',
        'AF': '所需信息',
        'AH': '机会窗口',
        'AJ': '所需设备',
    }

    results = []
    row = 7

    while True:
        if not ws[f'A{row}'].value:
            break

        item = {}
        for col, name in col_mapping.items():
            cell = ws[f'{col}{row}']
            value = cell.value
            if value and isinstance(value, str) and value.startswith('='):
                value = ""
            if value == 0:
                item[name] = "0"
            elif value:
                item[name] = value
            else:
                item[name] = ""

        results.append(item)
        row += 1

    return results

def process_file(excel_file):
    file_name = os.path.basename(excel_file).replace('.xlsx', '')

    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb[wb.sheetnames[-1]]
        tara_data = process_tara_data(ws)
        wb.close()

        if not tara_data:
            log_msg(f"[{file_name}] 无数据")
            return []

        interface_file = DATA_RAW_PATH / INTERFACE_FOLDER / "MY26.xlsx"
        topology_file = DATA_PROCESSED_PATH / TOPOLOGY_FOLDER / "MY26.txt"

        interface_data = []
        if interface_file.exists():
            log_msg(f"找到外部接口文件: {interface_file}")
            try:
                wb = openpyxl.load_workbook(str(interface_file))
                ws = wb.active
                row = 2
                while True:
                    info = ws[f'A{row}'].value
                    part = ws[f'B{row}'].value
                    if not info and not part:
                        break
                    interface_data.append({
                        "外部接口信息": info if info else "",
                        "外部接口关联部件": part if part else ""
                    })
                    row += 1
                wb.close()
                log_msg(f"读取外部接口数据: {len(interface_data)} 条")
            except Exception as e:
                log_msg(f"读取外部接口失败: {e}")
        else:
            log_msg(f"外部接口文件不存在: {interface_file}")

        topology_data = []
        if topology_file.exists():
            log_msg(f"找到拓扑图文件: {topology_file}")
            try:
                with open(str(topology_file), 'r', encoding='utf-8') as f:
                    for line in f:
                        line = line.strip()
                        if line:
                            topology_data.append(line)
                log_msg(f"读取拓扑图数据: {len(topology_data)} 条")
            except Exception as e:
                log_msg(f"读取拓扑图失败: {e}")
        else:
            log_msg(f"拓扑图文件不存在: {topology_file}")

        for item in tara_data:
            item["外部接口"] = interface_data
            item["拓扑图"] = topology_data

        log_msg(f"[{file_name}] 完成: {len(tara_data)} 条")
        return tara_data

    except Exception as e:
        log_msg(f"[{file_name}] 错误: {e}")
        import traceback
        traceback.print_exc()
        return []

def main():
    base_path = DATA_RAW_PATH / "MY26"
    all_data = []

    file_list = []
    for root, dirs, files in os.walk(str(base_path)):
        for f in files:
            if f.endswith('.xlsx') and not f.startswith('~$'):
                excel_file = os.path.join(root, f)
                file_list.append(excel_file)

    log_msg(f"找到 {len(file_list)} 个文件")

    for i, excel_file in enumerate(file_list):
        log_msg(f"[{i+1}/{len(file_list)}] 处理: {os.path.basename(excel_file)}")
        file_data = process_file(excel_file)
        all_data.extend(file_data)

    log_msg(f"总计: {len(all_data)} 条数据")

    output_dir = DATA_PROCESSED_PATH / "MY26"
    log_msg(f"输出目录: {output_dir}")
    output_dir.mkdir(parents=True, exist_ok=True)

    output_file = output_dir / "MY26.json"
    log_msg(f"输出文件: {output_file}")
    
    with open(str(output_file), 'w', encoding='utf-8') as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)

    log_msg(f"已保存到: {output_file}")

    with open('batch_log.txt', 'w', encoding='utf-8') as f:
        f.write('\n'.join(output_lines))

if __name__ == "__main__":
    main()